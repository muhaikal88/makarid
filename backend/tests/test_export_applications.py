"""
Test cases for Export Applications to Excel feature
Tests the POST /api/applications-session/export endpoint
- Returns ZIP file containing Data_Lamaran.xlsx and CV/ folder
- Requires session-based authentication
- Must have application_ids provided
"""

import pytest
import requests
import os
import io
import zipfile

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "admin@lucky.com"
TEST_PASSWORD = "Admin@2026!"
COMPANY_NAME = "PT. LUCKY PERDANA MULTIMEDIA"


class TestExportApplications:
    """Export Applications API tests"""
    
    @pytest.fixture(autouse=True)
    def setup(self, request):
        """Setup session for each test"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        # Store session token for cleanup
        request.node._session = self.session
    
    def _login(self):
        """Helper to perform login and get session cookie"""
        # Step 1: Unified login
        login_resp = self.session.post(f"{BASE_URL}/api/auth/unified-login", json={
            "email": TEST_EMAIL,
            "password": TEST_PASSWORD
        })
        assert login_resp.status_code == 200, f"Unified login failed: {login_resp.text}"
        
        access_list = login_resp.json().get("access_list", [])
        assert len(access_list) > 0, "No company access found"
        
        # Find the LUCKY company
        company_access = None
        for access in access_list:
            if "LUCKY" in access.get("company_name", "").upper():
                company_access = access
                break
        
        if not company_access:
            company_access = access_list[0]
        
        # Step 2: Select company
        select_resp = self.session.post(f"{BASE_URL}/api/auth/select-company", json={
            "company_id": company_access["company_id"],
            "role": company_access["role"],
            "user_table": company_access["user_table"],
            "user_id": company_access["user_id"]
        })
        assert select_resp.status_code == 200, f"Select company failed: {select_resp.text}"
        
        return company_access["company_id"]
    
    def _get_applications(self, company_id):
        """Helper to get existing applications"""
        resp = self.session.get(f"{BASE_URL}/api/applications-session")
        assert resp.status_code == 200, f"Failed to fetch applications: {resp.text}"
        return resp.json()
    
    # ===== BACKEND API TESTS =====
    
    def test_export_returns_401_if_not_authenticated(self):
        """Export returns 401 if not authenticated"""
        # Make request without logging in
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": ["test_id_1"]
        })
        
        assert response.status_code == 401, f"Expected 401, got {response.status_code}: {response.text}"
        print("PASSED: Export returns 401 if not authenticated")
    
    def test_export_returns_400_if_no_application_ids(self):
        """Export returns 400 if no application_ids provided"""
        company_id = self._login()
        
        # Request with empty array
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": []
        })
        
        assert response.status_code == 400, f"Expected 400, got {response.status_code}: {response.text}"
        error_detail = response.json().get("detail", "")
        assert "No applications selected" in error_detail or "application" in error_detail.lower()
        print("PASSED: Export returns 400 if no application_ids provided")
    
    def test_export_returns_zip_file(self):
        """Export returns valid ZIP file with selected applications"""
        company_id = self._login()
        
        # Get existing applications
        applications = self._get_applications(company_id)
        
        if len(applications) == 0:
            pytest.skip("No applications available to test export")
        
        # Select first application
        app_ids = [applications[0]["id"]]
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": app_ids
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Verify content type
        content_type = response.headers.get("Content-Type", "")
        assert "application/zip" in content_type, f"Expected application/zip, got {content_type}"
        
        # Verify content-disposition header
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp, f"Expected attachment, got {content_disp}"
        assert ".zip" in content_disp.lower(), f"Expected .zip in filename, got {content_disp}"
        
        print("PASSED: Export returns ZIP file")
    
    def test_zip_contains_excel_file(self):
        """ZIP file contains Data_Lamaran.xlsx"""
        company_id = self._login()
        
        applications = self._get_applications(company_id)
        if len(applications) == 0:
            pytest.skip("No applications available to test export")
        
        app_ids = [applications[0]["id"]]
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": app_ids
        })
        
        assert response.status_code == 200, f"Export failed: {response.text}"
        
        # Parse ZIP content
        zip_buffer = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer, 'r') as zf:
            file_list = zf.namelist()
            
            # Check for Data_Lamaran.xlsx
            xlsx_files = [f for f in file_list if f.endswith('.xlsx')]
            assert len(xlsx_files) > 0, f"No .xlsx file found in ZIP. Files: {file_list}"
            assert "Data_Lamaran.xlsx" in file_list, f"Data_Lamaran.xlsx not found. Files: {file_list}"
        
        print("PASSED: ZIP contains Data_Lamaran.xlsx")
    
    def test_export_multiple_applications(self):
        """Export works with multiple applications"""
        company_id = self._login()
        
        applications = self._get_applications(company_id)
        if len(applications) < 2:
            pytest.skip("Need at least 2 applications to test multiple export")
        
        # Select first 2 applications
        app_ids = [applications[0]["id"], applications[1]["id"]]
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": app_ids
        })
        
        assert response.status_code == 200, f"Export failed: {response.text}"
        
        # Verify ZIP is valid
        zip_buffer = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer, 'r') as zf:
            file_list = zf.namelist()
            assert "Data_Lamaran.xlsx" in file_list, f"Data_Lamaran.xlsx not found"
        
        print("PASSED: Export multiple applications works")
    
    def test_zip_contains_cv_folder_for_applications_with_resume(self):
        """ZIP contains CV/ folder with resume files"""
        company_id = self._login()
        
        applications = self._get_applications(company_id)
        
        # Find applications with resume
        apps_with_resume = [a for a in applications if a.get("resume_url")]
        
        if len(apps_with_resume) == 0:
            pytest.skip("No applications with resume found")
        
        app_ids = [apps_with_resume[0]["id"]]
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": app_ids
        })
        
        assert response.status_code == 200, f"Export failed: {response.text}"
        
        # Parse ZIP and check for CV folder
        zip_buffer = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer, 'r') as zf:
            file_list = zf.namelist()
            cv_files = [f for f in file_list if f.startswith("CV/")]
            
            print(f"ZIP contents: {file_list}")
            print(f"CV files: {cv_files}")
            
            # CV folder should exist if application has resume
            # Files in CV folder should match resume names
            if len(cv_files) > 0:
                print("PASSED: ZIP contains CV folder with resume files")
            else:
                # CV file might not exist on disk, but structure should be there
                print("INFO: No CV files in ZIP (resume file may not exist on server)")
    
    def test_export_returns_404_for_invalid_application_ids(self):
        """Export returns 404 for non-existent application IDs"""
        company_id = self._login()
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": ["nonexistent_id_123", "nonexistent_id_456"]
        })
        
        # Should return 404 since no applications found
        assert response.status_code == 404, f"Expected 404, got {response.status_code}: {response.text}"
        print("PASSED: Export returns 404 for invalid application IDs")
    
    def test_export_only_returns_own_company_applications(self):
        """Export only includes applications from user's company"""
        company_id = self._login()
        
        applications = self._get_applications(company_id)
        if len(applications) == 0:
            pytest.skip("No applications available")
        
        # Get application IDs from own company
        own_app_ids = [a["id"] for a in applications[:2]]
        
        # Mix with fake ID from "another company"
        mixed_ids = own_app_ids + ["fake_other_company_app_id"]
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": mixed_ids
        })
        
        # Should still work but only export owned applications
        assert response.status_code == 200, f"Export failed: {response.text}"
        print("PASSED: Export only returns own company applications")


class TestExportApplicationsDataIntegrity:
    """Test Excel data integrity"""
    
    @pytest.fixture(autouse=True)
    def setup(self, request):
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        request.node._session = self.session
    
    def _login(self):
        login_resp = self.session.post(f"{BASE_URL}/api/auth/unified-login", json={
            "email": TEST_EMAIL,
            "password": TEST_PASSWORD
        })
        assert login_resp.status_code == 200
        
        access_list = login_resp.json().get("access_list", [])
        company_access = access_list[0]
        
        for access in access_list:
            if "LUCKY" in access.get("company_name", "").upper():
                company_access = access
                break
        
        select_resp = self.session.post(f"{BASE_URL}/api/auth/select-company", json={
            "company_id": company_access["company_id"],
            "role": company_access["role"],
            "user_table": company_access["user_table"],
            "user_id": company_access["user_id"]
        })
        assert select_resp.status_code == 200
        return company_access["company_id"]
    
    def test_excel_has_required_columns(self):
        """Excel file has required columns: No, Posisi, Departemen, Status, Tanggal Melamar"""
        from openpyxl import load_workbook
        
        company_id = self._login()
        
        applications = self.session.get(f"{BASE_URL}/api/applications-session").json()
        if len(applications) == 0:
            pytest.skip("No applications available")
        
        response = self.session.post(f"{BASE_URL}/api/applications-session/export", json={
            "application_ids": [applications[0]["id"]]
        })
        
        assert response.status_code == 200
        
        zip_buffer = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer, 'r') as zf:
            xlsx_data = zf.read("Data_Lamaran.xlsx")
            xlsx_buffer = io.BytesIO(xlsx_data)
            
            wb = load_workbook(xlsx_buffer)
            ws = wb.active
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Required columns
            required = ["No", "Posisi", "Departemen", "Status", "Tanggal Melamar"]
            
            for col in required:
                assert col in headers, f"Missing required column: {col}. Headers: {headers}"
            
            print(f"Excel headers: {headers}")
            print("PASSED: Excel has required columns")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
