from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Request, Response, Cookie
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Any, Dict
import uuid
from datetime import datetime, timezone, timedelta
import hashlib
import jwt
import secrets
import json
import aiofiles
import httpx
import pyotp
import zipfile
import io
import tempfile
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create uploads directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', secrets.token_hex(32))
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app without a prefix
app = FastAPI(title="Makar.id HR System API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()
optional_security = HTTPBearer(auto_error=False)

# ============ MODELS ============

class UserRole:
    ADMIN = "admin"
    EMPLOYEE = "employee"

# ===== SUPER ADMIN MODELS (Separate Table) =====

class SuperAdminBase(BaseModel):
    email: EmailStr
    name: str
    picture: Optional[str] = None
    totp_secret: Optional[str] = None  # For Google Authenticator
    totp_enabled: bool = False
    is_active: bool = True

class SuperAdminCreate(BaseModel):
    email: EmailStr
    name: str
    password: str

class SuperAdminUpdate(BaseModel):
    email: Optional[EmailStr] = None
    name: Optional[str] = None
    password: Optional[str] = None
    is_active: Optional[bool] = None

class SuperAdmin(SuperAdminBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SuperAdminResponse(BaseModel):
    id: str
    email: str
    name: str
    is_active: bool
    created_at: str
    updated_at: str

class LoginResponseSuperAdmin(BaseModel):
    token: str
    user: dict  # Flexible dict to handle both SuperAdmin and User

# ===== COMPANY ADMIN MODELS (New Separate Table) =====

class CompanyAdminBase(BaseModel):
    email: EmailStr
    name: str
    picture: Optional[str] = None
    companies: List[str] = []  # List of company_ids user is admin of
    totp_secret: Optional[str] = None  # For Google Authenticator
    totp_enabled: bool = False
    is_active: bool = True
    auth_provider: str = "email"  # "email" or "google"

class CompanyAdminCreate(BaseModel):
    email: EmailStr
    name: str
    password: Optional[str] = None  # Optional if using Google OAuth
    company_id: str  # Initial company
    auth_provider: str = "email"

class CompanyAdmin(CompanyAdminBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"admin_{uuid.uuid4().hex[:12]}")
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ===== EMPLOYEE MODELS (New Separate Table) =====

class EmployeeBase(BaseModel):
    email: EmailStr
    name: str
    picture: Optional[str] = None
    companies: List[str] = []  # List of company_ids user is employee of
    is_active: bool = True
    auth_provider: str = "email"  # "email" or "google"

class EmployeeCreate(BaseModel):
    email: EmailStr
    name: str
    password: Optional[str] = None  # Optional if using Google OAuth
    company_id: str  # Initial company
    auth_provider: str = "email"

class Employee(EmployeeBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"emp_{uuid.uuid4().hex[:12]}")
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ===== UNIFIED LOGIN MODELS =====

class UnifiedLoginRequest(BaseModel):
    email: EmailStr
    password: str

class UserAccess(BaseModel):
    company_id: str
    company_name: str
    company_slug: str
    role: str  # "admin" or "employee"
    user_table: str  # "company_admins" or "employees"
    user_id: str

class UnifiedLoginResponse(BaseModel):
    access_list: List[UserAccess]
    user_email: str
    user_name: str
    user_picture: Optional[str] = None
    needs_selection: bool

class CompanyRoleSelection(BaseModel):
    company_id: str
    role: str
    user_table: str
    user_id: str

class SessionResponse(BaseModel):
    session_token: str
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    company_id: str
    company_name: str
    company_slug: str
    role: str

# ===== OLD USER MODELS (Keep for backward compatibility) =====

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str = UserRole.ADMIN  # Only 'admin' or 'employee'
    company_id: str  # REQUIRED - all users must belong to a company
    is_active: bool = True

class UserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    role: str = UserRole.ADMIN
    company_id: str  # REQUIRED

class UserUpdate(BaseModel):
    email: Optional[EmailStr] = None
    name: Optional[str] = None
    password: Optional[str] = None
    is_active: Optional[bool] = None
    role: Optional[str] = None

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    company_id: str
    is_active: bool
    created_at: str
    updated_at: str

class CompanyBase(BaseModel):
    name: str
    slug: str  # URL-friendly identifier (e.g., demo) - used for subdomain
    domain: str  # Primary identifier (e.g., company.co.id)
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[EmailStr] = None
    logo_url: Optional[str] = None
    is_active: bool = True
    # License management
    license_start: Optional[str] = None  # ISO date string
    license_end: Optional[str] = None    # ISO date string
    license_type: str = "trial"  # trial, monthly, yearly, lifetime
    # Custom domain settings for white-label (optional)
    custom_domains: Optional[Dict[str, str]] = None  # {"main": "company.co.id", "careers": "careers.company.co.id", "hr": "hr.company.co.id"}
    page_title: Optional[str] = None  # Custom HTML page title for white-label domains
    # SMTP settings (optional - for custom email notifications)
    smtp_settings: Optional[Dict[str, str]] = None  # {"host": "smtp.gmail.com", "port": "587", "user": "noreply@company.com", "password": "***", "from_email": "noreply@company.com", "from_name": "Company Name"}

class CompanyCreate(CompanyBase):
    pass

class CompanyUpdate(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    domain: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[EmailStr] = None
    logo_url: Optional[str] = None
    is_active: Optional[bool] = None
    license_start: Optional[str] = None
    license_end: Optional[str] = None
    license_type: Optional[str] = None
    custom_domains: Optional[Dict[str, str]] = None
    page_title: Optional[str] = None
    smtp_settings: Optional[Dict[str, str]] = None

class Company(CompanyBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompanyResponse(BaseModel):
    id: str
    name: str
    slug: Optional[str] = None
    domain: str
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    logo_url: Optional[str] = None
    is_active: bool
    license_start: Optional[str] = None
    license_end: Optional[str] = None
    license_type: Optional[str] = None
    license_status: Optional[str] = None  # active, expired, suspended
    days_remaining: Optional[int] = None
    created_at: str
    updated_at: str
    admin_count: int = 0
    employee_count: int = 0
    custom_domains: Optional[Dict[str, str]] = None
    custom_domains: Optional[Dict[str, str]] = None

# Company Profile Models
class CompanyProfileUpdate(BaseModel):
    tagline: Optional[str] = None
    description: Optional[str] = None
    vision: Optional[str] = None
    mission: Optional[str] = None
    history: Optional[str] = None
    culture: Optional[str] = None
    benefits: Optional[List[str]] = None
    social_links: Optional[Dict[str, str]] = None
    gallery_images: Optional[List[str]] = None
    cover_image: Optional[str] = None

class CompanyProfileResponse(BaseModel):
    id: str
    name: str
    domain: str
    logo_url: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    tagline: Optional[str] = None
    description: Optional[str] = None
    vision: Optional[str] = None
    mission: Optional[str] = None
    history: Optional[str] = None
    culture: Optional[str] = None
    benefits: Optional[List[str]] = None
    social_links: Optional[Dict[str, str]] = None
    gallery_images: Optional[List[str]] = None
    cover_image: Optional[str] = None

# ===== SYSTEM SETTINGS MODELS =====

class SMTPSettings(BaseModel):
    host: str
    port: int
    username: str
    password: str
    from_email: str
    from_name: str
    use_tls: bool = True

class SystemSettings(BaseModel):
    smtp_settings: Optional[SMTPSettings] = None

class SystemSettingsUpdate(BaseModel):
    smtp_settings: Optional[Dict[str, Any]] = None

# Job Posting Models
class JobStatus:
    DRAFT = "draft"
    PUBLISHED = "published"
    CLOSED = "closed"

class JobType:
    FULL_TIME = "full_time"
    PART_TIME = "part_time"
    CONTRACT = "contract"
    INTERNSHIP = "internship"

class JobCreate(BaseModel):
    title: str
    department: Optional[str] = None
    location: Optional[str] = None
    job_type: str = JobType.FULL_TIME
    description: str
    requirements: Optional[List[str]] = None
    responsibilities: Optional[List[str]] = None
    salary_min: Optional[int] = None
    salary_max: Optional[int] = None
    show_salary: bool = False
    status: str = JobStatus.DRAFT

class JobUpdate(BaseModel):
    title: Optional[str] = None
    department: Optional[str] = None
    location: Optional[str] = None
    job_type: Optional[str] = None
    description: Optional[str] = None
    requirements: Optional[List[str]] = None
    responsibilities: Optional[List[str]] = None
    salary_min: Optional[int] = None
    salary_max: Optional[int] = None
    show_salary: Optional[bool] = None
    status: Optional[str] = None

class JobResponse(BaseModel):
    id: str
    company_id: str
    title: str
    department: Optional[str] = None
    location: Optional[str] = None
    job_type: str
    description: str
    requirements: Optional[List[str]] = None
    responsibilities: Optional[List[str]] = None
    salary_min: Optional[int] = None
    salary_max: Optional[int] = None
    show_salary: bool
    status: str
    application_count: int = 0
    created_at: str
    updated_at: str

# Application Form Field Models
class FieldType:
    TEXT = "text"
    TEXTAREA = "textarea"
    EMAIL = "email"
    PHONE = "phone"
    NUMBER = "number"
    DATE = "date"
    SELECT = "select"
    CHECKBOX = "checkbox"
    FILE = "file"

class FormFieldCreate(BaseModel):
    field_name: str
    field_label: str
    field_type: str = FieldType.TEXT
    is_required: bool = True
    options: Optional[List[str]] = None  # For select fields
    placeholder: Optional[str] = None
    order: int = 0

class FormFieldUpdate(BaseModel):
    field_label: Optional[str] = None
    field_type: Optional[str] = None
    is_required: Optional[bool] = None
    options: Optional[List[str]] = None
    placeholder: Optional[str] = None
    order: Optional[int] = None

class FormFieldResponse(BaseModel):
    id: str
    company_id: str
    field_name: str
    field_label: str
    field_type: str
    is_required: bool
    options: Optional[List[str]] = None
    placeholder: Optional[str] = None
    order: int

# Application Models
class ApplicationStatus:
    PENDING = "pending"
    REVIEWING = "reviewing"
    SHORTLISTED = "shortlisted"
    INTERVIEWED = "interviewed"
    OFFERED = "offered"
    HIRED = "hired"
    REJECTED = "rejected"

class ApplicationCreate(BaseModel):
    job_id: str
    form_data: Dict[str, Any]

class ApplicationUpdateStatus(BaseModel):
    status: str
    notes: Optional[str] = None

class ApplicationResponse(BaseModel):
    id: str
    job_id: str
    company_id: str
    job_title: str
    job_department: Optional[str] = None
    applicant_name: str
    applicant_email: str
    form_data: Dict[str, Any]
    resume_url: Optional[str] = None
    status: str
    notes: Optional[str] = None
    created_at: str
    updated_at: str

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class LoginResponse(BaseModel):
    token: str
    user: UserResponse

class DashboardStats(BaseModel):
    total_companies: int
    active_companies: int
    total_users: int
    total_employees: int
    recent_companies: List[CompanyResponse]



# Activity Log Models
class ActivityLog(BaseModel):
    id: str
    user_id: str
    user_name: str
    user_email: str
    user_role: str  # "super_admin", "admin", "employee"
    company_id: Optional[str] = None
    company_name: Optional[str] = None
    action: str  # "create", "update", "delete", "login", "logout"
    resource_type: str  # "company", "user", "job", "application", "license", "profile", etc
    resource_id: Optional[str] = None
    description: str
    ip_address: Optional[str] = None
    timestamp: str


# ============ HELPERS ============

def validate_password_strength(password: str) -> tuple[bool, str]:
    """
    Validate password meets ISO 27001 standards:
    - Minimum 8 characters
    - At least 1 uppercase letter
    - At least 1 lowercase letter
    - At least 1 number
    - At least 1 special character
    """
    if len(password) < 8:
        return False, "Password must be at least 8 characters"
    
    if not re.search(r'[A-Z]', password):
        return False, "Password must contain at least 1 uppercase letter"
    
    if not re.search(r'[a-z]', password):
        return False, "Password must contain at least 1 lowercase letter"
    
    if not re.search(r'\d', password):
        return False, "Password must contain at least 1 number"
    
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False, "Password must contain at least 1 special character"
    
    return True, "Password is strong"

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed

def generate_slug(name: str) -> str:
    """Generate URL-friendly slug from company name"""
    import re
    # Convert to lowercase and replace spaces/special chars with hyphens
    slug = name.lower()
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    slug = slug.strip('-')
    return slug

def get_license_status(company: dict) -> tuple:
    """
    Returns (status, days_remaining)
    status: 'active', 'expired', 'suspended', 'no_license'
    """
    if not company.get("is_active"):
        return ("suspended", None)
    
    license_end = company.get("license_end")
    if not license_end:
        # No license set - treat as active (trial/lifetime)
        license_type = company.get("license_type", "trial")
        if license_type == "lifetime":
            return ("active", None)
        return ("active", None)  # Trial without end date
    
    try:
        end_date = datetime.fromisoformat(license_end.replace('Z', '+00:00'))
        now = datetime.now(timezone.utc)
        delta = end_date - now
        days_remaining = delta.days
        
        if days_remaining < 0:
            return ("expired", days_remaining)
        return ("active", days_remaining)
    except:
        return ("active", None)


async def create_activity_log(
    user_id: str,
    user_name: str, 
    user_email: str,
    user_role: str,
    action: str,
    resource_type: str,
    description: str,
    company_id: str = None,
    company_name: str = None,
    resource_id: str = None,
    ip_address: str = None
):
    """Helper to create activity log"""
    log_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "user_name": user_name,
        "user_email": user_email,
        "user_role": user_role,
        "company_id": company_id,
        "company_name": company_name,
        "action": action,
        "resource_type": resource_type,
        "resource_id": resource_id,
        "description": description,
        "ip_address": ip_address,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.activity_logs.insert_one(log_doc)


async def get_smtp_settings(company_id: str = None):
    """Get SMTP settings - company-specific first, fallback to global"""
    if company_id:
        company = await db.companies.find_one({"id": company_id}, {"_id": 0})
        if company and company.get("smtp_settings"):
            smtp = company["smtp_settings"]
            if smtp.get("host") and smtp.get("username") and smtp.get("password"):
                return smtp
    # Fallback to global
    settings = await db.system_settings.find_one({}, {"_id": 0})
    if settings and settings.get("smtp_settings"):
        smtp = settings["smtp_settings"]
        if smtp.get("host") and smtp.get("username") and smtp.get("password"):
            return smtp
    return None

async def send_notification_email(to_email: str, subject: str, html_body: str, text_body: str, company_id: str = None):
    """Send email notification using SMTP. Returns True on success, False on failure."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.utils import formatdate, make_msgid
    import asyncio
    
    smtp = await get_smtp_settings(company_id)
    if not smtp:
        logging.warning(f"SMTP not configured, skipping email to {to_email}")
        return False
    
    msg = MIMEMultipart("alternative")
    msg["From"] = f"{smtp.get('from_name', 'Makar.id')} <{smtp.get('from_email', smtp['username'])}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid(domain="makar.id")
    msg["Reply-To"] = smtp.get("from_email", smtp["username"])
    msg["X-Mailer"] = "Makar.id"
    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    
    def _send():
        port = int(smtp.get("port", 587))
        if port == 465:
            server = smtplib.SMTP_SSL(smtp["host"], port, timeout=20)
        else:
            server = smtplib.SMTP(smtp["host"], port, timeout=20)
            if smtp.get("use_tls", True):
                server.starttls()
        server.login(smtp["username"], smtp["password"])
        server.sendmail(smtp.get("from_email", smtp["username"]), to_email, msg.as_string())
        server.quit()
    
    try:
        # Run synchronous smtplib in thread pool to avoid blocking event loop
        await asyncio.to_thread(_send)
        logging.info(f"Email sent to {to_email}: {subject}")
        # Store success log
        await db.email_logs.insert_one({
            "to": to_email, "subject": subject, "status": "sent",
            "company_id": company_id,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        return True
    except Exception as e:
        error_msg = str(e)
        logging.error(f"Failed to send email to {to_email}: {error_msg}")
        # Store error log for visibility
        await db.email_logs.insert_one({
            "to": to_email, "subject": subject, "status": "failed",
            "error": error_msg, "company_id": company_id,
            "smtp_host": smtp.get("host"), "smtp_port": smtp.get("port"),
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        return False

STATUS_LABELS = {
    "pending": "Menunggu Review",
    "reviewing": "Sedang Direview",
    "shortlisted": "Masuk Shortlist",
    "interviewed": "Tahap Interview",
    "offered": "Mendapat Penawaran",
    "hired": "Diterima",
    "rejected": "Ditolak",
}

def build_email_header(company_name: str, company_logo: str = None):
    return f'''<div style="background:#2E4DA7;padding:28px 24px;text-align:center;border-radius:12px 12px 0 0;">
        <h1 style="color:#fff;margin:0;font-size:22px;">{company_name}</h1>
    </div>'''

def build_email_footer(company_name: str):
    return f'''<div style="background:#f9fafb;padding:14px 24px;text-align:center;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;">
        <p style="color:#9ca3af;font-size:12px;margin:0;">2026 {company_name} - Powered by Makar.id</p>
    </div>'''

async def send_application_confirmation_email(application: dict, job: dict, company: dict):
    """Send confirmation email when applicant submits application"""
    form_data = application.get("form_data", {})
    applicant_name = form_data.get("full_name", form_data.get("name", "Pelamar"))
    applicant_email = form_data.get("email")
    if not applicant_email:
        return
    
    company_name = company.get("name", "Perusahaan")
    job_title = job.get("title", "Posisi")
    
    subject = f"Konfirmasi Lamaran - {job_title} di {company_name}"
    
    html_body = f'''<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
        {build_email_header(company_name)}
        <div style="background:#fff;padding:28px 24px;border:1px solid #e5e7eb;border-top:none;">
            <h2 style="color:#1f2937;margin:0 0 12px;font-size:18px;">Halo {applicant_name},</h2>
            <p style="color:#4b5563;line-height:1.7;margin:0 0 16px;">
                Terima kasih telah melamar posisi <b>{job_title}</b> di <b>{company_name}</b>.
                Lamaran Anda telah kami terima dan sedang dalam proses peninjauan.
            </p>
            <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:16px;margin:16px 0;">
                <p style="color:#1e40af;margin:0 0 8px;font-size:14px;"><b>Detail Lamaran:</b></p>
                <p style="color:#1e40af;margin:0;font-size:13px;">Posisi: <b>{job_title}</b></p>
                <p style="color:#1e40af;margin:4px 0 0;font-size:13px;">Perusahaan: {company_name}</p>
                <p style="color:#1e40af;margin:4px 0 0;font-size:13px;">Status: Menunggu Review</p>
            </div>
            <p style="color:#6b7280;font-size:13px;line-height:1.6;">
                Kami akan menghubungi Anda jika ada perkembangan. Harap bersabar menunggu proses seleksi.
            </p>
        </div>
        {build_email_footer(company_name)}
    </div>'''
    
    text_body = f"Halo {applicant_name},\n\nTerima kasih telah melamar posisi {job_title} di {company_name}.\nLamaran Anda telah kami terima dan sedang dalam proses peninjauan.\n\nPosisi: {job_title}\nPerusahaan: {company_name}\nStatus: Menunggu Review\n\nKami akan menghubungi Anda jika ada perkembangan."
    
    await send_notification_email(applicant_email, subject, html_body, text_body, company.get("id"))

async def send_status_update_email(application: dict, job: dict, company: dict, old_status: str, new_status: str):
    """Send notification email when application status is updated"""
    form_data = application.get("form_data", {})
    applicant_name = form_data.get("full_name", form_data.get("name", "Pelamar"))
    applicant_email = form_data.get("email")
    if not applicant_email:
        return
    
    company_name = company.get("name", "Perusahaan")
    job_title = job.get("title", "Posisi")
    status_label = STATUS_LABELS.get(new_status, new_status.title())
    old_status_label = STATUS_LABELS.get(old_status, old_status.title())
    
    # Color based on status
    if new_status == "hired":
        status_bg, status_border, status_color = "#f0fdf4", "#bbf7d0", "#166534"
    elif new_status == "rejected":
        status_bg, status_border, status_color = "#fef2f2", "#fecaca", "#991b1b"
    elif new_status in ("offered", "shortlisted"):
        status_bg, status_border, status_color = "#eff6ff", "#bfdbfe", "#1e40af"
    else:
        status_bg, status_border, status_color = "#fffbeb", "#fde68a", "#92400e"
    
    if new_status in ('hired', 'offered'):
        closing_msg = 'Selamat! Tim kami akan segera menghubungi Anda untuk langkah selanjutnya.'
    elif new_status == 'rejected':
        closing_msg = 'Terima kasih telah melamar. Jangan berkecil hati, tetap semangat!'
    else:
        closing_msg = 'Terima kasih atas kesabaran Anda dalam proses seleksi ini.'
    
    subject = f"Info Lamaran - {job_title} di {company_name}"
    
    html_body = f'''<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
        {build_email_header(company_name)}
        <div style="background:#fff;padding:28px 24px;border:1px solid #e5e7eb;border-top:none;">
            <h2 style="color:#1f2937;margin:0 0 12px;font-size:18px;">Halo {applicant_name},</h2>
            <p style="color:#4b5563;line-height:1.7;margin:0 0 16px;">
                Ada pembaruan status untuk lamaran Anda di posisi <b>{job_title}</b>.
            </p>
            <div style="background:{status_bg};border:1px solid {status_border};border-radius:8px;padding:20px;margin:16px 0;text-align:center;">
                <p style="color:{status_color};margin:0;font-size:18px;font-weight:bold;">{status_label}</p>
            </div>
            <div style="margin:16px 0;font-size:13px;color:#4b5563;">
                <p style="margin:0 0 6px;"><b>Posisi:</b> {job_title}</p>
                <p style="margin:0 0 6px;"><b>Perusahaan:</b> {company_name}</p>
                <p style="margin:0;"><b>Status sebelumnya:</b> {old_status_label}</p>
            </div>
            <p style="color:#6b7280;font-size:13px;line-height:1.6;">{closing_msg}</p>
        </div>
        {build_email_footer(company_name)}
    </div>'''
    
    text_body = f"Halo {applicant_name},\n\nAda pembaruan status lamaran Anda:\n\nPosisi: {job_title}\nPerusahaan: {company_name}\nStatus Baru: {status_label}\nStatus Sebelumnya: {old_status_label}\n\n{closing_msg}"
    
    await send_notification_email(applicant_email, subject, html_body, text_body, company.get("id"))


    try:
        end_date = datetime.fromisoformat(license_end.replace('Z', '+00:00'))
        now = datetime.now(timezone.utc)
        delta = end_date - now
        days_remaining = delta.days
        
        if days_remaining < 0:
            return ("expired", days_remaining)
        return ("active", days_remaining)
    except:
        return ("active", None)

def build_company_response(company: dict, admin_count: int = 0, emp_count: int = 0) -> CompanyResponse:
    """Helper to build CompanyResponse with license status"""
    license_status, days_remaining = get_license_status(company)
    
    return CompanyResponse(
        id=company["id"],
        name=company["name"],
        slug=company.get("slug"),
        domain=company["domain"],
        address=company.get("address"),
        phone=company.get("phone"),
        email=company.get("email"),
        logo_url=company.get("logo_url"),
        is_active=company.get("is_active", True),
        license_start=company.get("license_start"),
        license_end=company.get("license_end"),
        license_type=company.get("license_type", "trial"),
        license_status=license_status,
        days_remaining=days_remaining,
        created_at=company["created_at"] if isinstance(company["created_at"], str) else company["created_at"].isoformat(),
        updated_at=company["updated_at"] if isinstance(company["updated_at"], str) else company["updated_at"].isoformat(),
        admin_count=admin_count,
        employee_count=emp_count,
        custom_domains=company.get("custom_domains")
    )

async def check_company_license(company_id: str = None, domain: str = None):
    """Check if company license is valid. Raises HTTPException if not."""
    if company_id:
        company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    elif domain:
        company = await db.companies.find_one({"domain": domain}, {"_id": 0})
    else:
        return None
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    status, _ = get_license_status(company)
    
    if status == "suspended":
        raise HTTPException(
            status_code=403, 
            detail="Company account is suspended. Please contact administrator."
        )
    elif status == "expired":
        raise HTTPException(
            status_code=403, 
            detail="Company license has expired. Please renew your subscription."
        )
    
    return company

def create_token(user_id: str, role: str, company_id: str = None) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "company_id": company_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        role = payload.get("role")
        
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # Check superadmins table first
        if role == "super_admin":
            user = await db.superadmins.find_one({"id": user_id}, {"_id": 0})
            if user:
                user["role"] = "super_admin"
                user["company_id"] = None
                return user
        
        # Check users table (company users)
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_optional_user(credentials: HTTPAuthorizationCredentials = Depends(optional_security)):
    if not credentials:
        return None
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        role = payload.get("role")
        
        if not user_id:
            return None
        
        # Check superadmins table first
        if role == "super_admin":
            user = await db.superadmins.find_one({"id": user_id}, {"_id": 0})
            if user:
                user["role"] = "super_admin"
                user["company_id"] = None
                return user
        
        # Check users table
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        return user
    except:
        return None

async def require_super_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super Admin access required")
    return current_user

async def require_admin_or_super(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["super_admin", UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# ============ STARTUP ============

@app.on_event("startup")
async def startup_event():
    """Auto-seed database from seed_data files if database is empty, otherwise create default superadmin."""
    total_admins = await db.superadmins.count_documents({})
    if total_admins == 0:
        # Try to seed from seed_data files first
        seed_dir = ROOT_DIR / "seed_data"
        if seed_dir.exists():
            seed_collections = [
                'superadmins', 'companies', 'company_admins', 'employees',
                'users', 'jobs', 'form_fields', 'applications', 'system_settings'
            ]
            seeded = False
            for coll_name in seed_collections:
                seed_file = seed_dir / f'{coll_name}.json'
                if seed_file.exists():
                    existing = await db[coll_name].count_documents({})
                    if existing == 0:
                        with open(seed_file) as f:
                            docs = json.load(f)
                        if docs:
                            await db[coll_name].insert_many(docs)
                            logging.info(f"Seeded {coll_name}: {len(docs)} documents")
                            seeded = True
            if seeded:
                logging.info("Database seeded from seed_data files")
                return
        
        # Fallback: create default superadmin
        super_admin = {
            "id": str(uuid.uuid4()),
            "email": "superadmin@makar.id",
            "name": "Super Admin",
            "password": hash_password("admin123"),
            "picture": None,
            "totp_secret": None,
            "totp_enabled": False,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.superadmins.insert_one(super_admin)
        logging.info("Default Super Admin created: superadmin@makar.id / admin123")
    else:
        logging.info(f"Super admins already exist: {total_admins} admin(s)")


# ============ AUTH ROUTES ============

# Super Admin Login (Separate endpoint)
class LoginWith2FA(BaseModel):
    email: EmailStr
    password: str
    totp_code: Optional[str] = None  # 6-digit code from Google Authenticator

@api_router.post("/auth/superadmin/login")
async def superadmin_login(data: LoginWith2FA):
    """Login for Super Admin with optional 2FA"""
    admin = await db.superadmins.find_one({"email": data.email}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(data.password, admin["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not admin.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is deactivated")
    
    # Check if 2FA is enabled
    if admin.get("totp_enabled") and admin.get("totp_secret"):
        if not data.totp_code:
            # Need 2FA code
            return {
                "requires_2fa": True,
                "message": "Please enter your 2FA code"
            }
        
        # Verify 2FA code
        totp = pyotp.TOTP(admin["totp_secret"])
        if not totp.verify(data.totp_code, valid_window=1):
            raise HTTPException(status_code=401, detail="Invalid 2FA code")
    
    token = create_token(admin["id"], "super_admin", None)
    
    # Log activity
    await create_activity_log(
        user_id=admin["id"],
        user_name=admin["name"],
        user_email=admin["email"],
        user_role="super_admin",
        action="login",
        resource_type="auth",
        description=f"Super Admin {admin['name']} logged in"
    )
    
    return {
        "token": token,
        "requires_2fa": False,
        "user": {
            "id": admin["id"],
            "email": admin["email"],
            "name": admin["name"],
            "picture": admin.get("picture"),
            "role": "super_admin",
            "company_id": None,
            "is_active": admin["is_active"],
            "created_at": admin["created_at"],
            "updated_at": admin["updated_at"]
        }
    }

# Company User Login (Separate endpoint)
@api_router.post("/auth/login", response_model=LoginResponse)
async def login(data: LoginRequest):
    """Login for Company Admin/Employee only"""
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is deactivated")
    
    # Check company license
    await check_company_license(company_id=user.get("company_id"))
    
    token = create_token(user["id"], user["role"], user.get("company_id"))
    
    return LoginResponse(
        token=token,
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            role=user["role"],
            company_id=user["company_id"],
            is_active=user["is_active"],
            created_at=user["created_at"] if isinstance(user["created_at"], str) else user["created_at"].isoformat(),
            updated_at=user["updated_at"] if isinstance(user["updated_at"], str) else user["updated_at"].isoformat()
        )
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return {
        "id": current_user["id"],
        "email": current_user["email"],
        "name": current_user["name"],
        "picture": current_user.get("picture"),
        "role": current_user["role"],
        "company_id": current_user.get("company_id"),
        "is_active": current_user["is_active"],
        "created_at": current_user["created_at"] if isinstance(current_user["created_at"], str) else current_user["created_at"].isoformat(),
        "updated_at": current_user["updated_at"] if isinstance(current_user["updated_at"], str) else current_user["updated_at"].isoformat()
    }


# ============ UNIFIED AUTH (Email-based across tables) ============

@api_router.post("/auth/unified-login", response_model=UnifiedLoginResponse)
async def unified_login(data: UnifiedLoginRequest):
    """
    Unified login for company admins & employees.
    Check email across both tables and return all access.
    """
    access_list = []
    user_name = None
    user_picture = None
    
    # Check company_admins table
    admin = await db.company_admins.find_one({"email": data.email}, {"_id": 0})
    if admin:
        if admin.get("password") and verify_password(data.password, admin["password"]):
            user_name = admin["name"]
            user_picture = admin.get("picture")
            
            # Get all companies this admin has access to
            for company_id in admin.get("companies", []):
                company = await db.companies.find_one({"id": company_id}, {"_id": 0})
                if company and company.get("is_active"):
                    # Check license
                    status, _ = get_license_status(company)
                    if status not in ["expired", "suspended"]:
                        access_list.append(UserAccess(
                            company_id=company_id,
                            company_name=company["name"],
                            company_slug=company.get("slug", company["domain"]),
                            role="admin",
                            user_table="company_admins",
                            user_id=admin["id"]
                        ))
    
    # Check employees table
    employee = await db.employees.find_one({"email": data.email}, {"_id": 0})
    if employee:
        if employee.get("password") and verify_password(data.password, employee["password"]):
            if not user_name:  # Use employee name if admin not found
                user_name = employee["name"]
                user_picture = employee.get("picture")
            
            # Get all companies this employee has access to
            for company_id in employee.get("companies", []):
                company = await db.companies.find_one({"id": company_id}, {"_id": 0})
                if company and company.get("is_active"):
                    status, _ = get_license_status(company)
                    if status not in ["expired", "suspended"]:
                        access_list.append(UserAccess(
                            company_id=company_id,
                            company_name=company["name"],
                            company_slug=company.get("slug", company["domain"]),
                            role="employee",
                            user_table="employees",
                            user_id=employee["id"]
                        ))
    
    if not access_list:
        raise HTTPException(status_code=401, detail="Invalid credentials or no active access")
    
    needs_selection = len(access_list) > 1
    
    return UnifiedLoginResponse(
        access_list=access_list,
        user_email=data.email,
        user_name=user_name,
        user_picture=user_picture,
        needs_selection=needs_selection
    )

@api_router.post("/auth/select-company", response_model=SessionResponse)
async def select_company(data: CompanyRoleSelection, response: Response):
    """
    After user selects company/role, create session and return session token
    """
    # Get user data based on table
    if data.user_table == "company_admins":
        user = await db.company_admins.find_one({"id": data.user_id}, {"_id": 0})
    else:
        user = await db.employees.find_one({"id": data.user_id}, {"_id": 0})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify user has access to this company
    if data.company_id not in user.get("companies", []):
        raise HTTPException(status_code=403, detail="No access to this company")
    
    # Get company info
    company = await db.companies.find_one({"id": data.company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Create session token
    session_token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session_doc = {
        "session_token": session_token,
        "user_id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "picture": user.get("picture"),
        "company_id": data.company_id,
        "company_name": company["name"],
        "role": data.role,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.user_sessions.insert_one(session_doc)
    
    # Log activity
    await create_activity_log(
        user_id=user["id"], user_name=user["name"], user_email=user["email"],
        user_role=data.role, action="login", resource_type="auth",
        description=f"{user['name']} login ke {company['name']} sebagai {data.role}",
        company_id=data.company_id, company_name=company["name"]
    )
    
    # Set httpOnly cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60,  # 7 days
        path="/"
    )
    
    return SessionResponse(
        session_token=session_token,
        user_id=user["id"],
        email=user["email"],
        name=user["name"],
        picture=user.get("picture"),
        company_id=data.company_id,
        company_name=company["name"],
        company_slug=company.get("slug"),
        role=data.role
    )

# ============ GOOGLE OAUTH ENDPOINTS ============

EMERGENT_AUTH_API = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

@api_router.post("/auth/google/callback")
async def google_oauth_callback(session_id: str, response: Response):
    """
    Process Google OAuth session_id from Emergent Auth.
    Return user's access list (companies & roles).
    """
    # Get user data from Emergent Auth
    async with httpx.AsyncClient() as client:
        try:
            resp = await client.get(
                EMERGENT_AUTH_API,
                headers={"X-Session-ID": session_id},
                timeout=10.0
            )
            resp.raise_for_status()
            google_data = resp.json()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to get session data: {str(e)}")
    
    email = google_data["email"]
    name = google_data["name"]
    picture = google_data.get("picture")
    
    access_list = []
    
    # Check company_admins table
    admin = await db.company_admins.find_one({"email": email}, {"_id": 0})
    if admin:
        # Update with Google data if needed
        await db.company_admins.update_one(
            {"id": admin["id"]},
            {"$set": {
                "name": name,
                "picture": picture,
                "auth_provider": "google",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Get access list
        for company_id in admin.get("companies", []):
            company = await db.companies.find_one({"id": company_id}, {"_id": 0})
            if company and company.get("is_active"):
                status, _ = get_license_status(company)
                if status not in ["expired", "suspended"]:
                    access_list.append({
                        "company_id": company_id,
                        "company_name": company["name"],
                        "company_slug": company.get("slug", company["domain"]),
                        "role": "admin",
                        "user_table": "company_admins",
                        "user_id": admin["id"]
                    })
    
    # Check employees table
    employee = await db.employees.find_one({"email": email}, {"_id": 0})
    if employee:
        # Update with Google data
        await db.employees.update_one(
            {"id": employee["id"]},
            {"$set": {
                "name": name,
                "picture": picture,
                "auth_provider": "google",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Get access list
        for company_id in employee.get("companies", []):
            company = await db.companies.find_one({"id": company_id}, {"_id": 0})
            if company and company.get("is_active"):
                status, _ = get_license_status(company)
                if status not in ["expired", "suspended"]:
                    access_list.append({
                        "company_id": company_id,
                        "company_name": company["name"],
                        "company_slug": company.get("slug", company["domain"]),
                        "role": "employee",
                        "user_table": "employees",
                        "user_id": employee["id"]
                    })
    
    if not access_list:
        raise HTTPException(
            status_code=403, 
            detail="No company access found. Please contact your administrator."
        )
    
    # If only 1 access, auto-create session
    if len(access_list) == 1:
        access = access_list[0]
        session_token = secrets.token_urlsafe(32)
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        session_doc = {
            "session_token": session_token,
            "user_id": access["user_id"],
            "email": email,
            "name": name,
            "picture": picture,
            "company_id": access["company_id"],
            "role": access["role"],
            "expires_at": expires_at.isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.user_sessions.insert_one(session_doc)
        
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            max_age=7*24*60*60,
            path="/"
        )
        
        company = await db.companies.find_one({"id": access["company_id"]}, {"_id": 0})
        
        return SessionResponse(
            session_token=session_token,
            user_id=access["user_id"],
            email=email,
            name=name,
            picture=picture,
            company_id=access["company_id"],
            company_name=company["name"],
            company_slug=company.get("slug"),
            role=access["role"]
        )
    
    # Multiple access - return list for selection
    return {
        "access_list": access_list,
        "user_email": email,
        "user_name": name,
        "user_picture": picture,
        "needs_selection": True
    }

# ============ SESSION-BASED AUTH HELPERS ============

async def get_session_user(request: Request):
    """Get user from session_token (cookie or header)"""
    # Try cookie first
    session_token = request.cookies.get("session_token")
    
    # Fallback to Authorization header
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Get session from database
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    # Check expiry
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        await db.user_sessions.delete_one({"session_token": session_token})
        raise HTTPException(status_code=401, detail="Session expired")
    
    return session

@api_router.get("/auth/me-session")
async def get_me_session(request: Request):
    """Get current user from session (cookie-based auth)"""
    session = await get_session_user(request)
    
    company = await db.companies.find_one({"id": session["company_id"]}, {"_id": 0})
    
    # Get custom domain for careers if set
    custom_careers_domain = None
    if company and company.get("custom_domains"):
        custom_careers_domain = company["custom_domains"].get("careers")
    
    return {
        "user_id": session["user_id"],
        "email": session["email"],
        "name": session["name"],
        "picture": session.get("picture"),
        "company_id": session["company_id"],
        "company_name": company["name"] if company else None,
        "company_slug": company.get("slug") if company else None,
        "custom_domain": custom_careers_domain,
        "role": session["role"]
    }

@api_router.post("/auth/logout")
async def logout_session(request: Request, response: Response):
    """Logout - delete session and clear cookie"""
    session_token = request.cookies.get("session_token")
    
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response.delete_cookie(
        key="session_token",
        path="/",
        httponly=True,
        secure=True,
        samesite="none"
    )


# ============ PROFILE MANAGEMENT ============

# Super Admin Profile
@api_router.get("/profile/superadmin")
async def get_superadmin_profile(current_user: dict = Depends(require_super_admin)):
    """Get super admin profile"""
    admin = await db.superadmins.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Profile not found")
    return admin

class SuperAdminProfileUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    picture: Optional[str] = None



# ============ SUPER ADMIN MANAGEMENT (Manage other super admins) ============

@api_router.get("/superadmins")
async def get_all_superadmins(current_user: dict = Depends(require_super_admin)):
    """Get list of all super admins"""
    admins = await db.superadmins.find({}, {"_id": 0, "password": 0}).to_list(100)
    return admins

class SuperAdminCreateRequest(BaseModel):
    email: EmailStr
    name: str
    password: str
    picture: Optional[str] = None

@api_router.post("/superadmins")
async def create_superadmin(data: SuperAdminCreateRequest, current_user: dict = Depends(require_super_admin)):
    """Create new super admin"""
    # Check if email exists
    existing = await db.superadmins.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Validate password
    is_valid, message = validate_password_strength(data.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=message)
    
    admin_doc = {
        "id": str(uuid.uuid4()),
        "email": data.email,
        "name": data.name,
        "password": hash_password(data.password),
        "picture": data.picture,
        "totp_secret": None,
        "totp_enabled": False,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.superadmins.insert_one(admin_doc)
    
    # Remove password from response
    admin_doc.pop("password")
    return admin_doc

class SuperAdminUpdateRequest(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    picture: Optional[str] = None
    is_active: Optional[bool] = None

@api_router.put("/superadmins/{admin_id}")
async def update_superadmin(admin_id: str, data: SuperAdminUpdateRequest, current_user: dict = Depends(require_super_admin)):
    """Update super admin (by another super admin)"""
    admin = await db.superadmins.find_one({"id": admin_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Super admin not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    if "password" in update_data:
        is_valid, message = validate_password_strength(update_data["password"])
        if not is_valid:
            raise HTTPException(status_code=400, detail=message)
        update_data["password"] = hash_password(update_data["password"])
    
    if "email" in update_data:
        existing = await db.superadmins.find_one({"email": update_data["email"], "id": {"$ne": admin_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.superadmins.update_one({"id": admin_id}, {"$set": update_data})
    
    updated = await db.superadmins.find_one({"id": admin_id}, {"_id": 0, "password": 0})
    return updated

@api_router.delete("/superadmins/{admin_id}")
async def delete_superadmin(admin_id: str, current_user: dict = Depends(require_super_admin)):
    """Delete super admin"""
    # Prevent self-deletion
    if admin_id == current_user["id"]:
        raise HTTPException(status_code=403, detail="Cannot delete your own account")
    
    admin = await db.superadmins.find_one({"id": admin_id})
    if not admin:
        raise HTTPException(status_code=404, detail="Super admin not found")
    
    # Check if this is the last admin
    total_admins = await db.superadmins.count_documents({})
    if total_admins <= 1:
        raise HTTPException(status_code=403, detail="Cannot delete the last super admin")
    
    await db.superadmins.delete_one({"id": admin_id})
    
    return {"message": "Super admin deleted successfully"}


# Toggle features for Super Admin management
@api_router.post("/superadmins/{admin_id}/toggle-2fa")
async def toggle_superadmin_2fa(admin_id: str, enable: bool, current_user: dict = Depends(require_super_admin)):
    """Enable or disable 2FA for another super admin"""
    admin = await db.superadmins.find_one({"id": admin_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Super admin not found")
    
    await db.superadmins.update_one(
        {"id": admin_id},
        {"$set": {
            "totp_enabled": False if not enable else admin.get("totp_enabled", False),
            "totp_secret": None if not enable else admin.get("totp_secret"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"2FA {'enabled' if enable else 'disabled'} successfully"}

@api_router.post("/superadmins/{admin_id}/toggle-active")
async def toggle_superadmin_active(admin_id: str, active: bool, current_user: dict = Depends(require_super_admin)):
    """Activate or deactivate super admin"""
    if admin_id == current_user["id"]:
        raise HTTPException(status_code=403, detail="Cannot deactivate your own account")
    
    await db.superadmins.update_one(
        {"id": admin_id},
        {"$set": {
            "is_active": active,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Account {'activated' if active else 'deactivated'} successfully"}

# Toggle for company admins and employees
@api_router.post("/users/{user_id}/toggle-2fa")
async def toggle_user_2fa(user_id: str, enable: bool, user_table: str, current_user: dict = Depends(require_super_admin)):
    """Enable or disable 2FA for company admin or employee"""
    if user_table == "company_admins":
        table = db.company_admins
    elif user_table == "employees":
        table = db.employees
    else:
        raise HTTPException(status_code=400, detail="Invalid user table")
    
    user = await table.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await table.update_one(
        {"id": user_id},
        {"$set": {
            "totp_enabled": False if not enable else user.get("totp_enabled", False),
            "totp_secret": None if not enable else user.get("totp_secret"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"2FA {'enabled' if enable else 'disabled'} successfully"}

@api_router.post("/users/{user_id}/toggle-active")
async def toggle_user_active(user_id: str, active: bool, user_table: str, current_user: dict = Depends(require_super_admin)):
    """Activate or deactivate company admin or employee"""
    if user_table == "company_admins":
        table = db.company_admins
    elif user_table == "employees":
        table = db.employees
    else:
        raise HTTPException(status_code=400, detail="Invalid user table")
    
    await table.update_one(
        {"id": user_id},
        {"$set": {
            "is_active": active,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Account {'activated' if active else 'deactivated'} successfully"}




@api_router.put("/profile/superadmin")
async def update_superadmin_profile(data: SuperAdminProfileUpdate, current_user: dict = Depends(require_super_admin)):
    """Update super admin profile"""
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    if "password" in update_data:
        # Validate password strength
        is_valid, message = validate_password_strength(update_data["password"])
        if not is_valid:
            raise HTTPException(status_code=400, detail=message)
        update_data["password"] = hash_password(update_data["password"])
    
    if "email" in update_data:
        # Check if email already used by another admin
        existing = await db.superadmins.find_one({"email": update_data["email"], "id": {"$ne": current_user["id"]}})
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.superadmins.update_one({"id": current_user["id"]}, {"$set": update_data})
    
    updated = await db.superadmins.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    return updated

# Session-based Profile (Company Admin & Employee)
@api_router.get("/profile/me")
async def get_my_profile(request: Request):
    """Get current user profile (session-based)"""
    session = await get_session_user(request)
    
    # Get user from appropriate table
    if session["role"] == "admin":
        user = await db.company_admins.find_one({"id": session["user_id"]}, {"_id": 0, "password": 0})
        table = "company_admins"
    elif session["role"] == "employee":
        user = await db.employees.find_one({"id": session["user_id"]}, {"_id": 0, "password": 0})
        table = "employees"
    else:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    if not user:
        raise HTTPException(status_code=404, detail="Profile not found")
    
    # Get company info for current session
    company = await db.companies.find_one({"id": session["company_id"]}, {"_id": 0})
    
    return {
        **user,
        "current_company": {
            "id": session["company_id"],
            "name": company["name"] if company else None,
            "slug": company.get("slug") if company else None
        },
        "current_role": session["role"],
        "user_table": table
    }


# ============ ACTIVITY LOGS ============

@api_router.get("/logs")
async def get_activity_logs(
    current_user: dict = Depends(require_super_admin),
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    company_id: Optional[str] = None,
    search: Optional[str] = None,
    start_date: Optional[str] = None,  # ISO format: 2026-01-01
    end_date: Optional[str] = None,    # ISO format: 2026-12-31
    limit: int = 100
):
    """Get activity logs (Super Admin only) with date range filter"""
    query = {}
    
    if user_id:
        query["user_id"] = user_id
    if action:
        query["action"] = action
    if resource_type:
        query["resource_type"] = resource_type
    if company_id:
        query["company_id"] = company_id
    if search:
        query["$or"] = [
            {"user_name": {"$regex": search, "$options": "i"}},
            {"user_email": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    # Date range filter
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            # Add one day to include end_date fully
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            date_query["$lt"] = end_dt.isoformat()
        query["timestamp"] = date_query
    
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    return logs

@api_router.get("/logs/me")
async def get_my_activity_logs(
    request: Request,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get activity logs for current company (Company Admin only) with pagination"""
    session = await require_session_admin(request)
    
    query = {"company_id": session["company_id"]}
    
    if action:
        query["action"] = action
    if resource_type:
        query["resource_type"] = resource_type
    if search:
        query["$or"] = [
            {"user_name": {"$regex": search, "$options": "i"}},
            {"user_email": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            date_query["$lt"] = end_dt.isoformat()
        query["timestamp"] = date_query
    
    total = await db.activity_logs.count_documents(query)
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "skip": skip, "limit": limit}

class ProfileUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    picture: Optional[str] = None

@api_router.put("/profile/me")
async def update_my_profile(data: ProfileUpdate, request: Request):
    """Update current user profile (session-based)"""
    session = await get_session_user(request)
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    if "password" in update_data:
        # Validate password strength
        is_valid, message = validate_password_strength(update_data["password"])
        if not is_valid:
            raise HTTPException(status_code=400, detail=message)
        update_data["password"] = hash_password(update_data["password"])
    
    # Determine table
    if session["role"] == "admin":
        table = db.company_admins
        table_name = "company_admins"
    elif session["role"] == "employee":
        table = db.employees
        table_name = "employees"
    else:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    if "email" in update_data:
        # Check if email already used
        existing_admin = await db.company_admins.find_one({"email": update_data["email"], "id": {"$ne": session["user_id"]}})
        existing_emp = await db.employees.find_one({"email": update_data["email"], "id": {"$ne": session["user_id"]}})
        if existing_admin or existing_emp:
            raise HTTPException(status_code=400, detail="Email already in use")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await table.update_one({"id": session["user_id"]}, {"$set": update_data})
    
    # Update session if name or email changed
    if "name" in update_data or "email" in update_data:
        session_update = {}
        if "name" in update_data:
            session_update["name"] = update_data["name"]
        if "email" in update_data:
            session_update["email"] = update_data["email"]
        await db.user_sessions.update_many(
            {"user_id": session["user_id"]},
            {"$set": session_update}
        )
    
    updated = await table.find_one({"id": session["user_id"]}, {"_id": 0, "password": 0})
    return updated


# ============ 2FA / TOTP ENDPOINTS ============

@api_router.post("/profile/superadmin/2fa/setup")
async def setup_superadmin_2fa(current_user: dict = Depends(require_super_admin)):
    """Setup Google Authenticator for super admin"""
    # Generate new TOTP secret
    secret = pyotp.random_base32()
    
    # Create provisioning URI for QR code
    totp = pyotp.TOTP(secret)
    provisioning_uri = totp.provisioning_uri(
        name=current_user["email"],
        issuer_name="Makar.id Super Admin"
    )
    
    # Save secret (but don't enable yet - will enable after verification)
    await db.superadmins.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "totp_secret": secret,
            "totp_enabled": False,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "secret": secret,
        "provisioning_uri": provisioning_uri,
        "qr_code_url": f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={provisioning_uri}"
    }

@api_router.post("/profile/superadmin/2fa/verify")
async def verify_superadmin_2fa(token: str, current_user: dict = Depends(require_super_admin)):
    """Verify and enable 2FA"""
    admin = await db.superadmins.find_one({"id": current_user["id"]}, {"_id": 0})
    
    if not admin or not admin.get("totp_secret"):
        raise HTTPException(status_code=400, detail="2FA not setup. Please setup first.")
    
    # Verify token
    totp = pyotp.TOTP(admin["totp_secret"])
    if not totp.verify(token, valid_window=1):
        raise HTTPException(status_code=400, detail="Invalid verification code")
    
    # Enable 2FA
    await db.superadmins.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "totp_enabled": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "2FA enabled successfully"}

@api_router.post("/profile/superadmin/2fa/disable")
async def disable_superadmin_2fa(password: str, current_user: dict = Depends(require_super_admin)):
    """Disable 2FA (requires password confirmation)"""
    admin = await db.superadmins.find_one({"id": current_user["id"]}, {"_id": 0})
    
    if not verify_password(password, admin["password"]):
        raise HTTPException(status_code=401, detail="Invalid password")
    
    await db.superadmins.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "totp_secret": None,
            "totp_enabled": False,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "2FA disabled successfully"}

# Company Admin 2FA
@api_router.post("/profile/me/2fa/setup")
async def setup_user_2fa(request: Request):
    """Setup Google Authenticator for company admin"""
    session = await get_session_user(request)
    
    if session["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only company admins can enable 2FA")
    
    # Generate secret
    secret = pyotp.random_base32()
    totp = pyotp.TOTP(secret)
    
    admin = await db.company_admins.find_one({"id": session["user_id"]}, {"_id": 0})
    provisioning_uri = totp.provisioning_uri(
        name=admin["email"],
        issuer_name="Makar.id Company Admin"
    )
    
    # Save secret
    await db.company_admins.update_one(
        {"id": session["user_id"]},
        {"$set": {
            "totp_secret": secret,
            "totp_enabled": False,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "secret": secret,
        "provisioning_uri": provisioning_uri,
        "qr_code_url": f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={provisioning_uri}"
    }

@api_router.post("/profile/me/2fa/verify")
async def verify_user_2fa(token: str, request: Request):
    """Verify and enable 2FA for company admin"""
    session = await get_session_user(request)
    
    if session["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only company admins can enable 2FA")
    
    admin = await db.company_admins.find_one({"id": session["user_id"]}, {"_id": 0})
    
    if not admin or not admin.get("totp_secret"):
        raise HTTPException(status_code=400, detail="2FA not setup")
    
    totp = pyotp.TOTP(admin["totp_secret"])
    if not totp.verify(token, valid_window=1):
        raise HTTPException(status_code=400, detail="Invalid verification code")
    
    await db.company_admins.update_one(
        {"id": session["user_id"]},
        {"$set": {
            "totp_enabled": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "2FA enabled successfully"}

@api_router.post("/profile/me/2fa/disable")
async def disable_user_2fa(password: str, request: Request):
    """Disable 2FA for company admin"""
    session = await get_session_user(request)
    
    if session["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only company admins can manage 2FA")
    
    admin = await db.company_admins.find_one({"id": session["user_id"]}, {"_id": 0})
    
    if not verify_password(password, admin["password"]):
        raise HTTPException(status_code=401, detail="Invalid password")
    
    await db.company_admins.update_one(
        {"id": session["user_id"]},
        {"$set": {
            "totp_secret": None,
            "totp_enabled": False,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "2FA disabled successfully"}


# ============ DASHBOARD ROUTES ============
    

async def require_session_admin(request: Request):
    """Require session-based admin authentication"""
    session = await get_session_user(request)
    if session["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return session

async def require_session_employee(request: Request):
    """Require session-based employee authentication"""
    session = await get_session_user(request)
    if session["role"] != "employee":
        raise HTTPException(status_code=403, detail="Employee access required")
    return session

async def require_session_user(request: Request):
    """Require any session-based authentication (admin or employee)"""
    session = await get_session_user(request)
    return session


    return {"message": "Logged out successfully"}


# ============ DASHBOARD ROUTES ============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(require_super_admin)):
    total_companies = await db.companies.count_documents({})
    active_companies = await db.companies.count_documents({"is_active": True})
    
    # Count from new tables
    total_admins = await db.company_admins.count_documents({})
    total_employees = await db.employees.count_documents({})
    total_users = total_admins + total_employees
    
    recent_companies_cursor = db.companies.find({}, {"_id": 0}).sort("created_at", -1).limit(5)
    recent_companies = await recent_companies_cursor.to_list(5)
    
    recent_with_counts = []
    for company in recent_companies:
        company_id = company["id"]
        admin_count = await db.company_admins.count_documents({"companies": company_id})
        emp_count = await db.employees.count_documents({"companies": company_id})
        recent_with_counts.append(build_company_response(company, admin_count, emp_count))
    
    return DashboardStats(
        total_companies=total_companies,
        active_companies=active_companies,
        total_users=total_users,
        total_employees=total_employees,
        recent_companies=recent_with_counts
    )


# ============ SYSTEM SETTINGS ROUTES (Super Admin Only) ============

@api_router.get("/system/settings")
async def get_system_settings(current_user: dict = Depends(require_super_admin)):
    """Get global system settings (SMTP, etc.)"""
    settings = await db.system_settings.find_one({}, {"_id": 0})
    
    if not settings:
        # Return default settings
        return {
            "smtp_settings": {
                "host": "",
                "port": 587,
                "username": "",
                "password": "",
                "from_email": "notif@makar.id",
                "from_name": "Makar.id Notifications",
                "use_tls": True
            }
        }
    
    return settings

@api_router.put("/system/settings")
async def update_system_settings(data: SystemSettingsUpdate, current_user: dict = Depends(require_super_admin)):
    """Update global system settings"""
    existing = await db.system_settings.find_one({})
    
    if existing:
        await db.system_settings.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                **data.model_dump(exclude_none=True),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        settings_doc = {
            **data.model_dump(exclude_none=True),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.system_settings.insert_one(settings_doc)
    
    return {"message": "System settings updated successfully"}


class TestEmailRequest(BaseModel):
    to_email: EmailStr

@api_router.post("/system/settings/test-email")
async def send_test_email(data: TestEmailRequest, current_user: dict = Depends(require_super_admin)):
    """Send a test email using the configured SMTP settings"""
    settings = await db.system_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("smtp_settings"):
        raise HTTPException(status_code=400, detail="SMTP belum dikonfigurasi. Silakan simpan pengaturan SMTP terlebih dahulu.")
    
    smtp = settings["smtp_settings"]
    
    if not smtp.get("host") or not smtp.get("username") or not smtp.get("password"):
        raise HTTPException(status_code=400, detail="Konfigurasi SMTP tidak lengkap (host, username, password wajib diisi).")
    
    subject = "Tes Email SMTP - Makar.id"
    
    html_body = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
    <div style="background:#2E4DA7;padding:28px 24px;text-align:center;border-radius:12px 12px 0 0;">
        <h1 style="color:#fff;margin:0;font-size:22px;">Makar.id</h1>
        <p style="color:#c7d2fe;margin:8px 0 0;font-size:14px;">Manajemen Karyawan</p>
    </div>
    <div style="background:#fff;padding:28px 24px;border:1px solid #e5e7eb;border-top:none;">
        <h2 style="color:#1f2937;margin:0 0 16px;font-size:18px;">Tes Email Berhasil!</h2>
        <p style="color:#4b5563;line-height:1.6;margin:0 0 16px;">
            Jika Anda menerima email ini, konfigurasi SMTP sudah benar dan siap digunakan.
        </p>
        <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:16px;margin:16px 0;">
            <p style="color:#166534;margin:0 0 8px;font-size:14px;"><b>Detail Konfigurasi:</b></p>
            <p style="color:#166534;margin:0;font-size:13px;">Host: {smtp.get('host')}</p>
            <p style="color:#166534;margin:4px 0 0;font-size:13px;">Port: {smtp.get('port')}</p>
            <p style="color:#166534;margin:4px 0 0;font-size:13px;">From: {smtp.get('from_email')}</p>
        </div>
        <p style="color:#6b7280;font-size:13px;margin:16px 0 0;">
            Dikirim oleh: {current_user.get('name', 'Super Admin')} ({current_user.get('email', '')})
        </p>
    </div>
    <div style="background:#f9fafb;padding:14px 24px;text-align:center;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;">
        <p style="color:#9ca3af;font-size:12px;margin:0;">2026 Makar.id - Manajemen Karyawan</p>
    </div>
    </div>"""
    
    text_body = f"Tes Email SMTP Makar.id\n\nJika Anda menerima email ini, konfigurasi SMTP sudah benar.\n\nHost: {smtp.get('host')}\nPort: {smtp.get('port')}\nFrom: {smtp.get('from_email')}\n\nDikirim oleh: {current_user.get('name', 'Super Admin')}"
    
    try:
        result = await send_notification_email(data.to_email, subject, html_body, text_body)
        
        if result:
            await create_activity_log(
                user_id=current_user["id"], user_name=current_user["name"], user_email=current_user["email"],
                user_role="super_admin", action="create", resource_type="email",
                description=f"Mengirim tes email ke {data.to_email}"
            )
            return {"message": f"Email tes berhasil dikirim ke {data.to_email}"}
        else:
            raise HTTPException(status_code=400, detail="Gagal mengirim email. Cek Log Pengiriman Email untuk detail error.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal mengirim email: {str(e)}")

@api_router.get("/system/email-logs")
async def get_email_logs(current_user: dict = Depends(require_super_admin), limit: int = 50):
    """Get email send logs for debugging"""
    logs = await db.email_logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    return logs



# ============ COMPANY ROUTES ============

@api_router.get("/companies", response_model=List[CompanyResponse])
async def get_companies(current_user: dict = Depends(require_super_admin)):
    companies = await db.companies.find({}, {"_id": 0}).to_list(1000)
    
    result = []
    for company in companies:
        company_id = company["id"]
        
        # Count admins yang punya akses ke company ini
        admin_count = await db.company_admins.count_documents({
            "companies": company_id
        })
        
        # Count employees yang punya akses ke company ini
        emp_count = await db.employees.count_documents({
            "companies": company_id
        })
        
        result.append(build_company_response(company, admin_count, emp_count))
    
    return result

@api_router.post("/companies", response_model=CompanyResponse)
async def create_company(data: CompanyCreate, current_user: dict = Depends(require_super_admin)):
    # Auto-generate slug if not provided
    if not hasattr(data, 'slug') or not data.slug:
        slug = generate_slug(data.name)
    else:
        slug = data.slug
    
    # Check if slug already exists
    existing_slug = await db.companies.find_one({"slug": slug})
    if existing_slug:
        # Append random number to make it unique
        import random
        slug = f"{slug}-{random.randint(1000, 9999)}"
    
    # Check if domain already exists
    existing = await db.companies.find_one({"domain": data.domain})
    if existing:
        raise HTTPException(status_code=400, detail="Domain already exists")
    
    company_dict = data.model_dump()
    company_dict["slug"] = slug
    company = Company(**company_dict)
    doc = company.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    
    # Set default license (30-day trial)
    doc["license_start"] = datetime.now(timezone.utc).isoformat()
    doc["license_end"] = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
    doc["license_type"] = data.license_type if hasattr(data, 'license_type') and data.license_type else "trial"
    
    # Initialize profile fields
    doc["profile"] = {
        "tagline": None,
        "description": None,
        "vision": None,
        "mission": None,
        "history": None,
        "culture": None,
        "benefits": [],
        "social_links": {},
        "gallery_images": [],
        "cover_image": None
    }
    
    # Initialize SMTP settings as null (will use default/system SMTP)
    doc["smtp_settings"] = None
    
    await db.companies.insert_one(doc)
    
    # Create default form fields for this company
    default_fields = [
        {"field_name": "full_name", "field_label": "Nama Lengkap", "field_type": "text", "is_required": True, "order": 1},
        {"field_name": "email", "field_label": "Email", "field_type": "email", "is_required": True, "order": 2},
        {"field_name": "phone", "field_label": "No. Telepon", "field_type": "phone", "is_required": True, "order": 3},
        {"field_name": "birth_place", "field_label": "Tempat Lahir", "field_type": "text", "is_required": True, "order": 4},
        {"field_name": "birth_date", "field_label": "Tanggal Lahir", "field_type": "date", "is_required": True, "order": 5},
        {"field_name": "education", "field_label": "Pendidikan Terakhir", "field_type": "select", "is_required": True, "order": 6, 
         "options": ["SD", "SMP", "SMA/SMK", "D3", "S1", "S2", "S3"]},
        {"field_name": "major", "field_label": "Jurusan", "field_type": "text", "is_required": True, "order": 7},
        {"field_name": "province", "field_label": "Provinsi", "field_type": "text", "is_required": True, "order": 8},
        {"field_name": "city", "field_label": "Kota/Kabupaten", "field_type": "text", "is_required": True, "order": 9},
        {"field_name": "district", "field_label": "Kecamatan", "field_type": "text", "is_required": True, "order": 10},
        {"field_name": "village", "field_label": "Kelurahan/Desa", "field_type": "text", "is_required": True, "order": 11},
        {"field_name": "full_address", "field_label": "Alamat Lengkap (RT/RW)", "field_type": "textarea", "is_required": True, "order": 12},
        {"field_name": "expected_salary", "field_label": "Gaji yang Diharapkan (Rp)", "field_type": "number", "is_required": True, "order": 13},
        {"field_name": "resume", "field_label": "Upload CV (PDF)", "field_type": "file", "is_required": True, "order": 14},
    ]
    
    for field in default_fields:
        field["id"] = str(uuid.uuid4())
        field["company_id"] = doc["id"]
        await db.form_fields.insert_one(field)
    
    return build_company_response(doc, 0)

@api_router.get("/companies/{company_id}", response_model=CompanyResponse)
async def get_company(company_id: str, current_user: dict = Depends(require_super_admin)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    admin_count = await db.company_admins.count_documents({"companies": company_id})
    emp_count = await db.employees.count_documents({"companies": company_id})
    
    return build_company_response(company, admin_count, emp_count)

@api_router.put("/companies/{company_id}", response_model=CompanyResponse)
async def update_company(company_id: str, data: CompanyUpdate, current_user: dict = Depends(require_super_admin)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.companies.update_one({"id": company_id}, {"$set": update_data})
    
    updated = await db.companies.find_one({"id": company_id}, {"_id": 0})
    admin_count = await db.company_admins.count_documents({"companies": company_id})
    emp_count = await db.employees.count_documents({"companies": company_id})
    
    return build_company_response(updated, admin_count, emp_count)

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str, current_user: dict = Depends(require_super_admin)):
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Delete all related data
    await db.company_admins.delete_many({"companies": company_id})
    await db.employees.delete_many({"companies": company_id})
    await db.jobs.delete_many({"company_id": company_id})
    await db.applications.delete_many({"company_id": company_id})
    await db.form_fields.delete_many({"company_id": company_id})
    await db.companies.delete_one({"id": company_id})
    
    return {"message": "Company deleted successfully"}

# ============ PUBLIC COMPANY LIST ============

@api_router.get("/public/companies")
async def get_public_companies():
    """Get list of active companies for login portal"""
    companies = await db.companies.find(
        {"is_active": True}, 
        {"_id": 0, "id": 1, "name": 1, "slug": 1, "domain": 1, "logo_url": 1}
    ).sort("name", 1).to_list(1000)
    
    return companies

# ============ DOMAIN LOOKUP ROUTES (For White-Label) ============

class DomainLookupResponse(BaseModel):
    found: bool
    company_id: Optional[str] = None
    company_name: Optional[str] = None
    slug: Optional[str] = None
    domain: Optional[str] = None
    page_type: Optional[str] = None  # "main", "careers", "hr"
    logo_url: Optional[str] = None

@api_router.get("/public/domain-lookup")
async def lookup_domain(hostname: str):
    """
    Lookup company by hostname for white-label routing.
    Checks custom_domains.main, custom_domains.careers, custom_domains.hr
    Also checks primary domain field and slug-based subdomain.
    """
    # Extract slug from subdomain if it's from makar.id
    # e.g., demo.makar.id -> slug=demo
    if '.makar.id' in hostname:
        slug = hostname.replace('.makar.id', '')
        company = await db.companies.find_one({
            "slug": slug,
            "is_active": True
        }, {"_id": 0})
        
        if company:
            return DomainLookupResponse(
                found=True,
                company_id=company["id"],
                company_name=company["name"],
                slug=company["slug"],
                domain=company["domain"],
                page_type="main",
                logo_url=company.get("logo_url")
            )
    
    # Check custom_domains or primary domain
    company = await db.companies.find_one({
        "$or": [
            {"custom_domains.main": hostname},
            {"custom_domains.careers": hostname},
            {"custom_domains.hr": hostname},
            {"domain": hostname}
        ],
        "is_active": True
    }, {"_id": 0})
    
    if not company:
        return DomainLookupResponse(found=False)
    
    # Determine page type
    custom_domains = company.get("custom_domains", {})
    page_type = "main"  # default
    
    if custom_domains:
        if custom_domains.get("careers") == hostname:
            page_type = "careers"
        elif custom_domains.get("hr") == hostname:
            page_type = "hr"
        elif custom_domains.get("main") == hostname:
            page_type = "main"
    
    return DomainLookupResponse(
        found=True,
        company_id=company["id"],
        company_name=company["name"],
        slug=company["slug"],
        domain=company["domain"],
        page_type=page_type,
        logo_url=company.get("logo_url")
    )

@api_router.put("/companies/{company_id}/domains")
async def update_company_domains(
    company_id: str, 
    domains: Dict[str, str],
    current_user: dict = Depends(require_super_admin)
):
    """
    Update custom domains for a company.
    domains: {"main": "luckycell.co.id", "careers": "careers.luckycell.co.id", "hr": "hr.luckycell.co.id"}
    """
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Validate domains are not used by other companies
    for domain_type, domain_value in domains.items():
        if domain_value:
            existing = await db.companies.find_one({
                "$or": [
                    {"custom_domains.main": domain_value},
                    {"custom_domains.careers": domain_value},
                    {"custom_domains.hr": domain_value},
                    {"domain": domain_value}
                ],
                "id": {"$ne": company_id}
            })
            if existing:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Domain {domain_value} is already used by another company"
                )
    
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {
            "custom_domains": domains,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Domains updated successfully", "domains": domains}

# ============ LICENSE MANAGEMENT ROUTES ============

class LicenseUpdate(BaseModel):
    license_start: Optional[str] = None
    license_end: Optional[str] = None
    license_type: Optional[str] = None
    is_active: Optional[bool] = None

@api_router.put("/companies/{company_id}/license")
async def update_company_license(
    company_id: str, 
    data: LicenseUpdate,
    current_user: dict = Depends(require_super_admin)
):
    """Update company license settings"""
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.companies.update_one({"id": company_id}, {"$set": update_data})
    
    updated = await db.companies.find_one({"id": company_id}, {"_id": 0})
    emp_count = await db.users.count_documents({"company_id": company_id})
    
    return build_company_response(updated, emp_count)

@api_router.post("/companies/{company_id}/license/extend")
async def extend_company_license(
    company_id: str,
    days: int = 30,
    current_user: dict = Depends(require_super_admin)
):
    """Extend company license by specified days"""
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Calculate new end date
    current_end = company.get("license_end")
    if current_end:
        try:
            end_date = datetime.fromisoformat(current_end.replace('Z', '+00:00'))
            # If expired, start from now
            if end_date < datetime.now(timezone.utc):
                end_date = datetime.now(timezone.utc)
        except:
            end_date = datetime.now(timezone.utc)
    else:
        end_date = datetime.now(timezone.utc)
    
    new_end = (end_date + timedelta(days=days)).isoformat()
    
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {
            "license_end": new_end,
            "is_active": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.companies.find_one({"id": company_id}, {"_id": 0})
    emp_count = await db.users.count_documents({"company_id": company_id})
    
    return {
        "message": f"License extended by {days} days",
        "new_end_date": new_end,
        "company": build_company_response(updated, emp_count)
    }

@api_router.post("/companies/{company_id}/suspend")
async def suspend_company(
    company_id: str,
    current_user: dict = Depends(require_super_admin)
):
    """Suspend a company (deactivate all access)"""
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {
            "is_active": False,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Company suspended successfully"}

@api_router.post("/companies/{company_id}/activate")
async def activate_company(
    company_id: str,
    current_user: dict = Depends(require_super_admin)
):
    """Activate a suspended company"""
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {
            "is_active": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Company activated successfully"}

# ============ PUBLIC COMPANY PROFILE ROUTES ============

@api_router.get("/public/company/{domain}", response_model=CompanyProfileResponse)
async def get_public_company_profile(domain: str):
    # Try to find by slug first, then domain
    company = await db.companies.find_one({
        "$or": [
            {"slug": domain},
            {"domain": domain}
        ]
    }, {"_id": 0})
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Check license status
    await check_company_license(company_id=company["id"])
    
    profile = company.get("profile", {})
    
    return CompanyProfileResponse(
        id=company["id"],
        name=company["name"],
        domain=company["domain"],
        logo_url=company.get("logo_url"),
        address=company.get("address"),
        phone=company.get("phone"),
        email=company.get("email"),
        tagline=profile.get("tagline"),
        description=profile.get("description"),
        vision=profile.get("vision"),
        mission=profile.get("mission"),
        history=profile.get("history"),
        culture=profile.get("culture"),
        benefits=profile.get("benefits"),
        social_links=profile.get("social_links"),
        gallery_images=profile.get("gallery_images"),
        cover_image=profile.get("cover_image")
    )

@api_router.put("/companies/{company_id}/profile", response_model=CompanyProfileResponse)
async def update_company_profile(company_id: str, data: CompanyProfileUpdate, current_user: dict = Depends(require_admin_or_super)):
    # Check access
    if current_user["role"] == UserRole.ADMIN and current_user.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    profile_update = {f"profile.{k}": v for k, v in data.model_dump().items() if v is not None}
    profile_update["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.companies.update_one({"id": company_id}, {"$set": profile_update})
    
    updated = await db.companies.find_one({"id": company_id}, {"_id": 0})
    profile = updated.get("profile", {})
    
    return CompanyProfileResponse(
        id=updated["id"],
        name=updated["name"],
        domain=updated["domain"],
        logo_url=updated.get("logo_url"),
        address=updated.get("address"),
        phone=updated.get("phone"),
        email=updated.get("email"),
        tagline=profile.get("tagline"),
        description=profile.get("description"),
        vision=profile.get("vision"),
        mission=profile.get("mission"),
        history=profile.get("history"),
        culture=profile.get("culture"),
        benefits=profile.get("benefits"),
        social_links=profile.get("social_links"),
        gallery_images=profile.get("gallery_images"),
        cover_image=profile.get("cover_image")
    )

# ============ COMPANY SETTINGS ROUTES (For Company Admin) ============

class CompanySettingsUpdate(BaseModel):
    custom_domains: Optional[Dict[str, str]] = None
    smtp_settings: Optional[Dict[str, str]] = None

@api_router.get("/company/settings")
async def get_company_settings(current_user: dict = Depends(require_admin_or_super)):
    """Get company settings (custom domains & SMTP) for Company Admin"""
    company_id = current_user.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="No company assigned")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    return {
        "id": company["id"],
        "name": company["name"],
        "slug": company.get("slug"),
        "domain": company["domain"],
        "custom_domains": company.get("custom_domains"),
        "smtp_settings": company.get("smtp_settings"),
        "default_subdomain": f"{company.get('slug')}.makar.id" if company.get('slug') else None
    }

@api_router.put("/company/settings")
async def update_company_settings(data: CompanySettingsUpdate, current_user: dict = Depends(require_admin_or_super)):
    """Update company settings (custom domains & SMTP) by Company Admin"""
    company_id = current_user.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="No company assigned")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    update_data = {}
    
    if data.custom_domains is not None:
        # Validate domains are not used by other companies
        for domain_type, domain_value in data.custom_domains.items():
            if domain_value:
                existing = await db.companies.find_one({
                    "$or": [
                        {"custom_domains.main": domain_value},
                        {"custom_domains.careers": domain_value},
                        {"custom_domains.hr": domain_value},
                        {"domain": domain_value}
                    ],
                    "id": {"$ne": company_id}
                })
                if existing:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Domain {domain_value} is already used by another company"
                    )
        update_data["custom_domains"] = data.custom_domains
    
    if data.smtp_settings is not None:
        update_data["smtp_settings"] = data.smtp_settings
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.companies.update_one({"id": company_id}, {"$set": update_data})
    
    return {"message": "Settings updated successfully", "updated_fields": list(update_data.keys())}

# ============ JOB POSTING ROUTES ============

@api_router.get("/jobs", response_model=List[JobResponse])
async def get_jobs(current_user: dict = Depends(require_admin_or_super)):
    query = {}
    if current_user["role"] == UserRole.ADMIN:
        query["company_id"] = current_user.get("company_id")
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    result = []
    for job in jobs:
        app_count = await db.applications.count_documents({"job_id": job["id"]})
        result.append(JobResponse(
            id=job["id"],
            company_id=job["company_id"],
            title=job["title"],
            department=job.get("department"),
            location=job.get("location"),
            job_type=job["job_type"],
            description=job["description"],
            requirements=job.get("requirements"),
            responsibilities=job.get("responsibilities"),
            salary_min=job.get("salary_min"),
            salary_max=job.get("salary_max"),
            show_salary=job.get("show_salary", False),
            status=job["status"],
            application_count=app_count,
            created_at=job["created_at"],
            updated_at=job["updated_at"]
        ))
    

    
    return result

# Session-based job endpoints (for new auth system)
@api_router.get("/jobs-session", response_model=List[JobResponse])
async def get_jobs_session(request: Request):
    """Get jobs using session auth"""
    session = await require_session_admin(request)
    
    query = {"company_id": session["company_id"]}
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    result = []
    for job in jobs:
        app_count = await db.applications.count_documents({"job_id": job["id"]})
        result.append(JobResponse(
            id=job["id"],
            company_id=job["company_id"],
            title=job["title"],
            department=job.get("department"),
            location=job.get("location"),
            job_type=job["job_type"],
            description=job["description"],
            requirements=job.get("requirements"),
            responsibilities=job.get("responsibilities"),
            salary_min=job.get("salary_min"),
            salary_max=job.get("salary_max"),
            show_salary=job.get("show_salary", False),
            status=job["status"],
            application_count=app_count,
            created_at=job["created_at"],
            updated_at=job["updated_at"]
        ))

    return result


@api_router.post("/jobs-session", response_model=JobResponse)
async def create_job_session(data: JobCreate, request: Request):
    """Create job using session auth"""
    session = await require_session_admin(request)
    
    job_doc = {
        "id": str(uuid.uuid4()),
        "company_id": session["company_id"],
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.jobs.insert_one(job_doc)
    
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="create", resource_type="job",
        description=f"Membuat lowongan baru: {job_doc['title']}",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=job_doc["id"]
    )
    
    return JobResponse(
        id=job_doc["id"],
        company_id=job_doc["company_id"],
        title=job_doc["title"],
        department=job_doc.get("department"),
        location=job_doc.get("location"),
        job_type=job_doc["job_type"],
        description=job_doc["description"],
        requirements=job_doc.get("requirements"),
        responsibilities=job_doc.get("responsibilities"),
        salary_min=job_doc.get("salary_min"),
        salary_max=job_doc.get("salary_max"),
        show_salary=job_doc.get("show_salary", False),
        status=job_doc["status"],
        application_count=0,
        created_at=job_doc["created_at"],
        updated_at=job_doc["updated_at"]
    )

@api_router.put("/jobs-session/{job_id}", response_model=JobResponse)
async def update_job_session(job_id: str, data: JobUpdate, request: Request):
    """Update job using session auth"""
    session = await require_session_admin(request)
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job["company_id"] != session["company_id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.jobs.update_one({"id": job_id}, {"$set": update_data})
    
    changes = ", ".join(update_data.keys())
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="update", resource_type="job",
        description=f"Update lowongan '{job['title']}': {changes}",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=job_id
    )
    
    updated = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    app_count = await db.applications.count_documents({"job_id": job_id})
    
    return JobResponse(
        id=updated["id"],
        company_id=updated["company_id"],
        title=updated["title"],
        department=updated.get("department"),
        location=updated.get("location"),
        job_type=updated["job_type"],
        description=updated["description"],
        requirements=updated.get("requirements"),
        responsibilities=updated.get("responsibilities"),
        salary_min=updated.get("salary_min"),
        salary_max=updated.get("salary_max"),
        show_salary=updated.get("show_salary", False),
        status=updated["status"],
        application_count=app_count,
        created_at=updated["created_at"],
        updated_at=updated["updated_at"]
    )


@api_router.delete("/jobs-session/{job_id}")
async def delete_job_session(job_id: str, request: Request):
    """Delete job using session auth"""
    session = await require_session_admin(request)
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job["company_id"] != session["company_id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Delete related applications
    await db.applications.delete_many({"job_id": job_id})
    await db.jobs.delete_one({"id": job_id})
    
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="delete", resource_type="job",
        description=f"Menghapus lowongan: {job['title']}",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=job_id
    )
    
    return {"message": "Job deleted successfully"}





@api_router.post("/jobs", response_model=JobResponse)
async def create_job(data: JobCreate, current_user: dict = Depends(require_admin_or_super)):
    company_id = current_user.get("company_id")
    if current_user["role"] == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=400, detail="Super Admin must specify company")
    
    if not company_id:
        raise HTTPException(status_code=400, detail="No company assigned")
    
    job_doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.jobs.insert_one(job_doc)
    
    return JobResponse(
        id=job_doc["id"],
        company_id=job_doc["company_id"],
        title=job_doc["title"],
        department=job_doc.get("department"),
        location=job_doc.get("location"),
        job_type=job_doc["job_type"],
        description=job_doc["description"],
        requirements=job_doc.get("requirements"),
        responsibilities=job_doc.get("responsibilities"),
        salary_min=job_doc.get("salary_min"),
        salary_max=job_doc.get("salary_max"),
        show_salary=job_doc.get("show_salary", False),
        status=job_doc["status"],
        application_count=0,
        created_at=job_doc["created_at"],
        updated_at=job_doc["updated_at"]
    )

@api_router.put("/jobs/{job_id}", response_model=JobResponse)
async def update_job(job_id: str, data: JobUpdate, current_user: dict = Depends(require_admin_or_super)):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if current_user["role"] == UserRole.ADMIN and job["company_id"] != current_user.get("company_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.jobs.update_one({"id": job_id}, {"$set": update_data})
    
    updated = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    app_count = await db.applications.count_documents({"job_id": job_id})
    
    return JobResponse(
        id=updated["id"],
        company_id=updated["company_id"],
        title=updated["title"],
        department=updated.get("department"),
        location=updated.get("location"),
        job_type=updated["job_type"],
        description=updated["description"],
        requirements=updated.get("requirements"),
        responsibilities=updated.get("responsibilities"),
        salary_min=updated.get("salary_min"),
        salary_max=updated.get("salary_max"),
        show_salary=updated.get("show_salary", False),
        status=updated["status"],
        application_count=app_count,
        created_at=updated["created_at"],
        updated_at=updated["updated_at"]
    )

@api_router.delete("/jobs/{job_id}")
async def delete_job(job_id: str, current_user: dict = Depends(require_admin_or_super)):
    job = await db.jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if current_user["role"] == UserRole.ADMIN and job["company_id"] != current_user.get("company_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.applications.delete_many({"job_id": job_id})
    await db.jobs.delete_one({"id": job_id})
    
    return {"message": "Job deleted successfully"}


# ============ WILAYAH.ID PROXY ROUTES (CORS workaround) ============

@api_router.get("/wilayah/provinces")
async def get_provinces():
    """Proxy for wilayah.id provinces API to bypass CORS"""
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get("https://wilayah.id/api/provinces.json", timeout=10.0)
            return response.json()
        except Exception as e:
            logging.error(f"Failed to fetch provinces: {e}")
            raise HTTPException(status_code=502, detail="Failed to fetch provinces from wilayah.id")

@api_router.get("/wilayah/regencies/{province_code}")
async def get_regencies(province_code: str):
    """Proxy for wilayah.id regencies/cities API to bypass CORS"""
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(f"https://wilayah.id/api/regencies/{province_code}.json", timeout=10.0)
            return response.json()
        except Exception as e:
            logging.error(f"Failed to fetch regencies: {e}")
            raise HTTPException(status_code=502, detail="Failed to fetch cities from wilayah.id")

@api_router.get("/wilayah/districts/{regency_code}")
async def get_districts(regency_code: str):
    """Proxy for wilayah.id districts API to bypass CORS"""
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(f"https://wilayah.id/api/districts/{regency_code}.json", timeout=10.0)
            return response.json()
        except Exception as e:
            logging.error(f"Failed to fetch districts: {e}")
            raise HTTPException(status_code=502, detail="Failed to fetch districts from wilayah.id")

@api_router.get("/wilayah/villages/{district_code}")
async def get_villages(district_code: str):
    """Proxy for wilayah.id villages API to bypass CORS"""
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(f"https://wilayah.id/api/villages/{district_code}.json", timeout=10.0)
            return response.json()
        except Exception as e:
            logging.error(f"Failed to fetch villages: {e}")
            raise HTTPException(status_code=502, detail="Failed to fetch villages from wilayah.id")


# ============ PUBLIC JOB ROUTES ============

@api_router.get("/public/careers/{domain}/jobs")
async def get_public_jobs(domain: str):
    # Try to find by slug first, then domain
    company = await db.companies.find_one({
        "$or": [
            {"slug": domain},
            {"domain": domain}
        ]
    }, {"_id": 0})
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Check license status
    await check_company_license(company_id=company["id"])
    
    jobs = await db.jobs.find(
        {"company_id": company["id"], "status": {"$in": ["published", "closed"]}}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    result = []
    for job in jobs:
        job_data = {
            "id": job["id"],
            "title": job["title"],
            "department": job.get("department"),
            "location": job.get("location"),
            "job_type": job["job_type"],
            "description": job["description"],
            "requirements": job.get("requirements"),
            "responsibilities": job.get("responsibilities"),
            "status": job["status"],
            "created_at": job["created_at"]
        }
        if job.get("show_salary"):
            job_data["salary_min"] = job.get("salary_min")
            job_data["salary_max"] = job.get("salary_max")
        result.append(job_data)
    
    return {
        "company": {
            "id": company["id"],
            "name": company["name"],
            "domain": company["domain"],
            "logo_url": company.get("logo_url"),
            "tagline": company.get("profile", {}).get("tagline"),
            "culture": company.get("profile", {}).get("culture"),
            "benefits": company.get("profile", {}).get("benefits")
        },
        "jobs": result
    }

@api_router.get("/public/careers/{domain}/jobs/{job_id}")
async def get_public_job_detail(domain: str, job_id: str):
    # Try to find by slug first, then domain
    company = await db.companies.find_one({
        "$or": [
            {"slug": domain},
            {"domain": domain}
        ]
    }, {"_id": 0})
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Check license status
    await check_company_license(company_id=company["id"])
    
    job = await db.jobs.find_one(
        {"id": job_id, "company_id": company["id"], "status": JobStatus.PUBLISHED}, 
        {"_id": 0}
    )
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get form fields
    form_fields = await db.form_fields.find(
        {"company_id": company["id"]}, 
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    job_data = {
        "id": job["id"],
        "title": job["title"],
        "department": job.get("department"),
        "location": job.get("location"),
        "job_type": job["job_type"],
        "description": job["description"],
        "requirements": job.get("requirements"),
        "responsibilities": job.get("responsibilities"),
        "created_at": job["created_at"]
    }
    if job.get("show_salary"):
        job_data["salary_min"] = job.get("salary_min")
        job_data["salary_max"] = job.get("salary_max")
    
    return {
        "company": {
            "id": company["id"],
            "name": company["name"],
            "domain": company["domain"],
            "logo_url": company.get("logo_url")
        },
        "job": job_data,
        "form_fields": form_fields
    }

# ============ FORM FIELDS ROUTES ============

@api_router.get("/form-fields", response_model=List[FormFieldResponse])
async def get_form_fields(current_user: dict = Depends(require_admin_or_super)):
    company_id = current_user.get("company_id")
    if current_user["role"] == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=400, detail="Super Admin must specify company")
    
    fields = await db.form_fields.find({"company_id": company_id}, {"_id": 0}).sort("order", 1).to_list(100)
    
    return [FormFieldResponse(**field) for field in fields]

@api_router.post("/form-fields", response_model=FormFieldResponse)
async def create_form_field(data: FormFieldCreate, current_user: dict = Depends(require_admin_or_super)):
    company_id = current_user.get("company_id")
    if current_user["role"] == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=400, detail="Super Admin must specify company")
    
    field_doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        **data.model_dump()
    }
    
    await db.form_fields.insert_one(field_doc)
    
    return FormFieldResponse(**field_doc)

@api_router.put("/form-fields/{field_id}", response_model=FormFieldResponse)
async def update_form_field(field_id: str, data: FormFieldUpdate, current_user: dict = Depends(require_admin_or_super)):
    field = await db.form_fields.find_one({"id": field_id}, {"_id": 0})
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    if current_user["role"] == UserRole.ADMIN and field["company_id"] != current_user.get("company_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    await db.form_fields.update_one({"id": field_id}, {"$set": update_data})
    
    updated = await db.form_fields.find_one({"id": field_id}, {"_id": 0})
    return FormFieldResponse(**updated)

@api_router.delete("/form-fields/{field_id}")
async def delete_form_field(field_id: str, current_user: dict = Depends(require_admin_or_super)):
    field = await db.form_fields.find_one({"id": field_id})
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    if current_user["role"] == UserRole.ADMIN and field["company_id"] != current_user.get("company_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.form_fields.delete_one({"id": field_id})
    return {"message": "Field deleted successfully"}

# ============ APPLICATION ROUTES ============

@api_router.post("/public/apply")
async def submit_application(
    job_id: str = Form(...),
    form_data: str = Form(...),
    resume: Optional[UploadFile] = File(None)
):
    job = await db.jobs.find_one({"id": job_id, "status": JobStatus.PUBLISHED}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    company = await db.companies.find_one({"id": job["company_id"]}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    try:
        parsed_data = json.loads(form_data)
    except:
        raise HTTPException(status_code=400, detail="Invalid form data")
    
    resume_url = None
    if resume:
        file_ext = resume.filename.split(".")[-1] if "." in resume.filename else "pdf"
        content = await resume.read()
        
        # Auto-compress images
        IMAGE_EXTS = {"jpg", "jpeg", "png", "webp", "bmp", "gif"}
        if file_ext.lower() in IMAGE_EXTS:
            from PIL import Image as PILImage
            img = PILImage.open(io.BytesIO(content))
            
            # Resize if too large (max 1920px width)
            max_width = 1920
            if img.width > max_width:
                ratio = max_width / img.width
                img = img.resize((max_width, int(img.height * ratio)), PILImage.LANCZOS)
            
            # Convert to JPEG
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            
            buf = io.BytesIO()
            img.save(buf, "JPEG", quality=75, optimize=True)
            content = buf.getvalue()
            file_ext = "jpg"
            logging.info(f"Compressed image: {resume.filename} -> {len(content)} bytes")
        
        file_name = f"{uuid.uuid4()}.{file_ext}"
        file_path = UPLOAD_DIR / file_name
        
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(content)
        
        resume_url = f"/api/uploads/{file_name}"
    
    application_doc = {
        "id": str(uuid.uuid4()),
        "job_id": job_id,
        "company_id": job["company_id"],
        "form_data": parsed_data,
        "resume_url": resume_url,
        "status": ApplicationStatus.PENDING,
        "notes": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.applications.insert_one(application_doc)
    
    # Send confirmation email to applicant (async, don't block response)
    try:
        await send_application_confirmation_email(application_doc, job, company)
    except Exception as e:
        logging.error(f"Failed to send confirmation email: {e}")
    
    return {"message": "Application submitted successfully", "id": application_doc["id"]}

@api_router.get("/applications", response_model=List[ApplicationResponse])
async def get_applications(
    job_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_admin_or_super)
):
    query = {}
    if current_user["role"] == UserRole.ADMIN:
        query["company_id"] = current_user.get("company_id")
    if job_id:
        query["job_id"] = job_id
    if status:
        query["status"] = status
    
    applications = await db.applications.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    result = []
    for app in applications:
        job = await db.jobs.find_one({"id": app["job_id"]}, {"_id": 0})
        job_title = job["title"] if job else "Unknown"
        job_department = job.get("department") if job else None
        
        form_data = app.get("form_data", {})
        applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
        applicant_email = form_data.get("email", "Unknown")
        
        result.append(ApplicationResponse(
            id=app["id"],
            job_id=app["job_id"],
            company_id=app["company_id"],
            job_title=job_title,
            job_department=job_department,
            applicant_name=applicant_name,
            applicant_email=applicant_email,
            form_data=form_data,
            resume_url=app.get("resume_url"),
            status=app["status"],
            notes=app.get("notes"),
            created_at=app["created_at"],
            updated_at=app["updated_at"]
        ))
    


@api_router.get("/applications-session", response_model=List[ApplicationResponse])
async def get_applications_session(
    request: Request,
    job_id: Optional[str] = None,
    status: Optional[str] = None
):
    """Get applications using session auth"""
    session = await require_session_admin(request)
    
    query = {"company_id": session["company_id"], "deleted_at": {"$exists": False}}
    if job_id:
        query["job_id"] = job_id
    if status:
        query["status"] = status
    
    applications = await db.applications.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    result = []
    for app in applications:
        job = await db.jobs.find_one({"id": app["job_id"]}, {"_id": 0})
        job_title = job["title"] if job else "Unknown"
        job_department = job.get("department") if job else None
        
        form_data = app.get("form_data", {})
        applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
        applicant_email = form_data.get("email", "Unknown")
        
        result.append(ApplicationResponse(
            id=app["id"],
            job_id=app["job_id"],
            company_id=app["company_id"],
            job_title=job_title,
            job_department=job_department,
            applicant_name=applicant_name,
            applicant_email=applicant_email,
            form_data=form_data,
            resume_url=app.get("resume_url"),
            status=app["status"],
            notes=app.get("notes"),
            created_at=app["created_at"],
            updated_at=app["updated_at"]
        ))
    
    return result

@api_router.put("/applications-session/{app_id}/status")
async def update_application_status_session(app_id: str, status: str, notes: Optional[str] = None, request: Request = None):
    """Update application status using session auth"""
    session = await require_session_admin(request)
    
    application = await db.applications.find_one({"id": app_id}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if application["company_id"] != session["company_id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    old_status = application.get("status", "unknown")
    await db.applications.update_one(
        {"id": app_id},
        {"$set": {
            "status": status,
            "notes": notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    form_data = application.get("form_data", {})
    applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="update", resource_type="application",
        description=f"Update status lamaran '{applicant_name}': {old_status} -> {status}",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=app_id
    )
    
    # Send email notification to applicant
    try:
        job = await db.jobs.find_one({"id": application["job_id"]}, {"_id": 0})
        company = await db.companies.find_one({"id": application["company_id"]}, {"_id": 0})
        if job and company:
            await send_status_update_email(application, job, company, old_status, status)
    except Exception as e:
        logging.error(f"Failed to send status update email: {e}")
    
    return {"message": "Status updated successfully"}

@api_router.get("/applications-session/{app_id}")
async def get_application_detail_session(app_id: str, request: Request):
    """Get application detail using session auth"""
    session = await require_session_admin(request)
    
    application = await db.applications.find_one({"id": app_id}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if application["company_id"] != session["company_id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get job info
    job = await db.jobs.find_one({"id": application["job_id"]}, {"_id": 0})
    
    return {
        **application,
        "job_title": job["title"] if job else "Unknown",
        "job": job
    }


@api_router.delete("/applications-session/{app_id}")
async def soft_delete_application(app_id: str, request: Request):
    """Soft delete - move to trash"""
    session = await require_session_admin(request)
    application = await db.applications.find_one({"id": app_id, "company_id": session["company_id"]}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    await db.applications.update_one({"id": app_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    
    form_data = application.get("form_data", {})
    applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="delete", resource_type="application",
        description=f"Memindahkan lamaran '{applicant_name}' ke tempat sampah",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=app_id
    )
    
    return {"message": "Application moved to trash"}

@api_router.get("/applications-session-trash", response_model=List[ApplicationResponse])
async def get_trash_applications(request: Request):
    """Get trashed applications"""
    session = await require_session_admin(request)
    applications = await db.applications.find(
        {"company_id": session["company_id"], "deleted_at": {"$exists": True}},
        {"_id": 0}
    ).sort("deleted_at", -1).to_list(1000)
    result = []
    for app in applications:
        job = await db.jobs.find_one({"id": app["job_id"]}, {"_id": 0})
        job_title = job["title"] if job else "Unknown"
        job_department = job.get("department") if job else None
        form_data = app.get("form_data", {})
        result.append(ApplicationResponse(
            id=app["id"], job_id=app["job_id"], company_id=app["company_id"],
            job_title=job_title, job_department=job_department,
            applicant_name=form_data.get("full_name", form_data.get("name", "Unknown")),
            applicant_email=form_data.get("email", "Unknown"),
            form_data=form_data, resume_url=app.get("resume_url"),
            status=app["status"], notes=app.get("notes"),
            created_at=app["created_at"], updated_at=app["updated_at"]
        ))
    return result

@api_router.post("/applications-session/{app_id}/restore")
async def restore_application(app_id: str, request: Request):
    """Restore from trash"""
    session = await require_session_admin(request)
    application = await db.applications.find_one({"id": app_id, "company_id": session["company_id"], "deleted_at": {"$exists": True}}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found in trash")
    await db.applications.update_one({"id": app_id}, {"$unset": {"deleted_at": ""}})
    
    form_data = application.get("form_data", {})
    applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="update", resource_type="application",
        description=f"Memulihkan lamaran '{applicant_name}' dari tempat sampah",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=app_id
    )
    
    return {"message": "Application restored"}

@api_router.delete("/applications-session/{app_id}/permanent")
async def permanent_delete_application(app_id: str, request: Request):
    """Permanent delete from trash"""
    session = await require_session_admin(request)
    application = await db.applications.find_one({"id": app_id, "company_id": session["company_id"], "deleted_at": {"$exists": True}}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found in trash")
    
    form_data = application.get("form_data", {})
    applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
    
    await db.applications.delete_one({"id": app_id})
    
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="delete", resource_type="application",
        description=f"Menghapus permanen lamaran: {applicant_name}",
        company_id=session["company_id"], company_name=session.get("company_name"),
        resource_id=app_id
    )
    
    return {"message": "Application permanently deleted"}


class ExportRequest(BaseModel):
    application_ids: List[str]

@api_router.post("/applications-session/export")
async def export_applications(data: ExportRequest, request: Request):
    """Export selected applications to Excel + CV files as ZIP"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    
    session = await require_session_admin(request)
    
    if not data.application_ids:
        raise HTTPException(status_code=400, detail="No applications selected")
    
    # Fetch applications
    apps = await db.applications.find(
        {"id": {"$in": data.application_ids}, "company_id": session["company_id"]},
        {"_id": 0}
    ).to_list(len(data.application_ids))
    
    if not apps:
        raise HTTPException(status_code=404, detail="No applications found")
    
    # Collect all unique form_data keys across all applications
    all_keys = []
    key_set = set()
    for app in apps:
        for key in app.get("form_data", {}).keys():
            if key not in key_set:
                all_keys.append(key)
                key_set.add(key)
    
    # Fetch job titles
    job_ids = list(set(a["job_id"] for a in apps))
    jobs_map = {}
    for jid in job_ids:
        job = await db.jobs.find_one({"id": jid}, {"_id": 0})
        if job:
            jobs_map[jid] = job
    
    IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
    THUMB_HEIGHT = 120  # pixels
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Lamaran"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E4DA7", end_color="2E4DA7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    link_font = Font(color="2E4DA7", underline="single", size=10)
    
    # Build headers
    headers = ["No", "Posisi", "Departemen", "Status", "Tanggal Melamar"]
    field_labels = {
        "full_name": "Nama Lengkap", "name": "Nama", "email": "Email",
        "phone": "No. Telepon", "tempat_lahir": "Tempat Lahir",
        "tanggal_lahir": "Tanggal Lahir", "gender": "Jenis Kelamin",
        "pendidikan": "Pendidikan", "jurusan": "Jurusan",
        "pengalaman": "Pengalaman Kerja", "alamat": "Alamat",
        "provinsi": "Provinsi", "kota": "Kota/Kabupaten",
        "kecamatan": "Kecamatan", "kelurahan": "Kelurahan",
        "address": "Alamat", "experience": "Pengalaman",
        "education": "Pendidikan", "cover_letter": "Surat Lamaran",
    }
    for key in all_keys:
        headers.append(field_labels.get(key, key.replace("_", " ").title()))
    headers.append("Preview CV")
    headers.append("File CV")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Track CV files to include in ZIP
    cv_files = []
    # Track temp thumbnail files to cleanup
    temp_thumbs = []
    
    cv_preview_col = 6 + len(all_keys)    # "Preview CV" column
    cv_link_col = cv_preview_col + 1       # "File CV" column
    
    # Set preview column width
    ws.column_dimensions[get_column_letter(cv_preview_col)].width = 22
    
    # Write data rows
    for row_idx, app in enumerate(apps, 2):
        form_data = app.get("form_data", {})
        job = jobs_map.get(app["job_id"], {})
        applicant_name = form_data.get("full_name", form_data.get("name", "Unknown"))
        
        # Fixed columns
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=job.get("title", "-")).border = thin_border
        ws.cell(row=row_idx, column=3, value=job.get("department", "-")).border = thin_border
        ws.cell(row=row_idx, column=4, value=app.get("status", "-")).border = thin_border
        
        date_str = app.get("created_at", "")
        try:
            dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
            date_str = dt.strftime("%d %b %Y %H:%M")
        except:
            pass
        ws.cell(row=row_idx, column=5, value=date_str).border = thin_border
        
        # Form data columns
        for col_offset, key in enumerate(all_keys):
            val = form_data.get(key, "")
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=6 + col_offset, value=str(val) if val else "")
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
        
        # CV columns
        resume_url = app.get("resume_url")
        if resume_url:
            filename = resume_url.split("/")[-1]
            file_path = UPLOAD_DIR / filename
            safe_name = f"{applicant_name.replace(' ', '_')}_{filename}"
            ext = Path(filename).suffix.lower()
            
            if file_path.exists():
                cv_files.append({"path": file_path, "name": f"CV/{safe_name}"})
                
                # Preview column: embed image or show link
                if ext in IMAGE_EXTS:
                    try:
                        # Create thumbnail
                        with PILImage.open(file_path) as img:
                            ratio = THUMB_HEIGHT / img.height
                            thumb_w = int(img.width * ratio)
                            thumb = img.resize((thumb_w, THUMB_HEIGHT), PILImage.LANCZOS)
                            if thumb.mode in ("RGBA", "P"):
                                thumb = thumb.convert("RGB")
                            
                            thumb_path = Path(tempfile.mktemp(suffix=".jpg"))
                            thumb.save(thumb_path, "JPEG", quality=80)
                            temp_thumbs.append(thumb_path)
                        
                        xl_img = XlImage(str(thumb_path))
                        cell_ref = f"{get_column_letter(cv_preview_col)}{row_idx}"
                        ws.add_image(xl_img, cell_ref)
                        
                        # Set row height to fit image
                        ws.row_dimensions[row_idx].height = THUMB_HEIGHT * 0.75 + 10
                    except Exception as e:
                        ws.cell(row=row_idx, column=cv_preview_col, value=f"(gagal preview: {ext})").border = thin_border
                else:
                    # Non-image file: show file type info
                    preview_cell = ws.cell(row=row_idx, column=cv_preview_col, value=f"[File {ext.upper()}]")
                    preview_cell.border = thin_border
                    preview_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Link column: hyperlink to CV file in ZIP
                link_cell = ws.cell(row=row_idx, column=cv_link_col, value=safe_name)
                link_cell.hyperlink = f"CV/{safe_name}"
                link_cell.font = link_font
                link_cell.border = thin_border
            else:
                ws.cell(row=row_idx, column=cv_preview_col, value="(file tidak ditemukan)").border = thin_border
                ws.cell(row=row_idx, column=cv_link_col, value="-").border = thin_border
        else:
            ws.cell(row=row_idx, column=cv_preview_col, value="-").border = thin_border
            ws.cell(row=row_idx, column=cv_link_col, value="-").border = thin_border
    
    # Auto-fit column widths (skip preview column which is fixed)
    for col_idx in range(1, len(headers) + 1):
        if col_idx == cv_preview_col:
            continue
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=len(apps) + 1):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    
    # Create ZIP in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Add Excel file
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        zf.writestr("Data_Lamaran.xlsx", excel_buffer.read())
        
        # Add CV files
        for cv in cv_files:
            zf.write(cv["path"], cv["name"])
    
    zip_buffer.seek(0)
    
    # Cleanup temp thumbnails
    for tp in temp_thumbs:
        try:
            tp.unlink()
        except:
            pass
    
    # Log activity
    await create_activity_log(
        user_id=session["user_id"], user_name=session["name"], user_email=session["email"],
        user_role="admin", action="create", resource_type="export",
        description=f"Export {len(apps)} data lamaran ke Excel",
        company_id=session["company_id"], company_name=session.get("company_name")
    )
    
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="Export_Lamaran_{timestamp}.zip"'}
    )


@api_router.put("/applications/{app_id}/status")
async def update_application_status(app_id: str, data: ApplicationUpdateStatus, current_user: dict = Depends(require_admin_or_super)):
    application = await db.applications.find_one({"id": app_id}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if current_user["role"] == UserRole.ADMIN and application["company_id"] != current_user.get("company_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_data = {
        "status": data.status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if data.notes is not None:
        update_data["notes"] = data.notes
    
    await db.applications.update_one({"id": app_id}, {"$set": update_data})
    
    return {"message": "Application status updated"}

@api_router.delete("/applications/{app_id}")
async def delete_application(app_id: str, current_user: dict = Depends(require_admin_or_super)):
    application = await db.applications.find_one({"id": app_id})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if current_user["role"] == UserRole.ADMIN and application["company_id"] != current_user.get("company_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.applications.delete_one({"id": app_id})
    return {"message": "Application deleted"}

# ============ FILE UPLOAD ROUTES ============

@api_router.get("/uploads/{filename}")
async def get_uploaded_file(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path)

# ============ USER ROUTES (Updated for new structure) ============

class AllUsersResponse(BaseModel):
    """Combined response for company admins and employees"""
    id: str
    email: str
    name: str
    role: str  # "admin" or "employee"
    companies: List[str]  # List of company IDs
    company_names: List[str]  # List of company names (for display)
    is_active: bool
    auth_provider: str
    created_at: str
    updated_at: str

@api_router.get("/users")
async def get_users(current_user: dict = Depends(require_super_admin)):
    """Get all company admins and employees"""
    result = []
    
    # Get all company admins
    admins = await db.company_admins.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for admin in admins:
        # Get company names
        company_names = []
        for comp_id in admin.get("companies", []):
            comp = await db.companies.find_one({"id": comp_id}, {"_id": 0, "name": 1})
            if comp:
                company_names.append(comp["name"])
        
        result.append({
            "id": admin["id"],
            "email": admin["email"],
            "name": admin["name"],
            "role": "admin",
            "companies": admin.get("companies", []),
            "company_names": company_names,
            "is_active": admin.get("is_active", True),
            "auth_provider": admin.get("auth_provider", "email"),
            "created_at": admin["created_at"] if isinstance(admin["created_at"], str) else admin["created_at"].isoformat(),
            "updated_at": admin["updated_at"] if isinstance(admin["updated_at"], str) else admin["updated_at"].isoformat()
        })
    
    # Get all employees
    employees = await db.employees.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for emp in employees:
        # Get company names
        company_names = []
        for comp_id in emp.get("companies", []):
            comp = await db.companies.find_one({"id": comp_id}, {"_id": 0, "name": 1})
            if comp:
                company_names.append(comp["name"])
        
        result.append({
            "id": emp["id"],
            "email": emp["email"],
            "name": emp["name"],
            "role": "employee",
            "companies": emp.get("companies", []),
            "company_names": company_names,
            "is_active": emp.get("is_active", True),
            "auth_provider": emp.get("auth_provider", "email"),
            "created_at": emp["created_at"] if isinstance(emp["created_at"], str) else emp["created_at"].isoformat(),
            "updated_at": emp["updated_at"] if isinstance(emp["updated_at"], str) else emp["updated_at"].isoformat()
        })
    
    return result

@api_router.post("/users", response_model=UserResponse)
async def create_user(data: UserCreate, current_user: dict = Depends(require_super_admin)):
    # Check if email already exists in both tables
    existing_admin = await db.company_admins.find_one({"email": data.email})
    existing_emp = await db.employees.find_one({"email": data.email})
    if existing_admin or existing_emp:
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    
    now = datetime.now(timezone.utc).isoformat()
    
    if data.role == "admin":
        user_id = f"admin_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "id": user_id,
            "email": data.email,
            "name": data.name,
            "password": hash_password(data.password),
            "companies": [data.company_id],
            "is_active": True,
            "auth_provider": "email",
            "created_at": now,
            "updated_at": now
        }
        await db.company_admins.insert_one(user_doc)
    else:
        user_id = f"emp_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "id": user_id,
            "email": data.email,
            "name": data.name,
            "password": hash_password(data.password),
            "companies": [data.company_id],
            "is_active": True,
            "auth_provider": "email",
            "created_at": now,
            "updated_at": now
        }
        await db.employees.insert_one(user_doc)
    
    return UserResponse(
        id=user_id,
        email=data.email,
        name=data.name,
        role=data.role,
        company_id=data.company_id,
        is_active=True,
        created_at=now,
        updated_at=now
    )

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, data: UserUpdate, current_user: dict = Depends(require_super_admin)):
    # Find in company_admins or employees
    user = await db.company_admins.find_one({"id": user_id}, {"_id": 0})
    table = db.company_admins
    role = "admin"
    if not user:
        user = await db.employees.find_one({"id": user_id}, {"_id": 0})
        table = db.employees
        role = "employee"
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if "password" in update_data:
        update_data["password"] = hash_password(update_data["password"])
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await table.update_one({"id": user_id}, {"$set": update_data})
    updated = await table.find_one({"id": user_id}, {"_id": 0})
    
    company_id = updated.get("companies", [None])[0] if updated.get("companies") else updated.get("company_id", "")
    
    return UserResponse(
        id=updated["id"],
        email=updated["email"],
        name=updated["name"],
        role=data.role or role,
        company_id=company_id,
        is_active=updated.get("is_active", True),
        created_at=updated["created_at"] if isinstance(updated["created_at"], str) else updated["created_at"].isoformat(),
        updated_at=updated["updated_at"] if isinstance(updated["updated_at"], str) else updated["updated_at"].isoformat()
    )

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_super_admin)):
    # Check company_admins first, then employees
    user = await db.company_admins.find_one({"id": user_id})
    if user:
        await db.company_admins.delete_one({"id": user_id})
        return {"message": "User deleted successfully"}
    
    user = await db.employees.find_one({"id": user_id})
    if user:
        await db.employees.delete_one({"id": user_id})
        return {"message": "User deleted successfully"}
    
    raise HTTPException(status_code=404, detail="User not found")

# ============ HEALTH CHECK ============

@api_router.get("/")
async def root():
    return {"message": "Lucky Cell HR System API", "status": "running"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# ============ DATABASE SEED ENDPOINT ============

class SeedRequest(BaseModel):
    superadmin_email: Optional[str] = "superadmin@makar.id"
    superadmin_password: Optional[str] = "admin123"
    superadmin_name: Optional[str] = "Super Admin"

@api_router.post("/seed/init")
async def seed_database(data: SeedRequest = SeedRequest()):
    """Initialize database with default superadmin. Only works if superadmins collection is empty."""
    total_admins = await db.superadmins.count_documents({})
    if total_admins > 0:
        return {
            "status": "skipped",
            "message": f"Database already has {total_admins} superadmin(s). Seed not needed.",
            "existing_admins": total_admins
        }
    
    super_admin = {
        "id": str(uuid.uuid4()),
        "email": data.superadmin_email,
        "name": data.superadmin_name,
        "password": hash_password(data.superadmin_password),
        "picture": None,
        "totp_secret": None,
        "totp_enabled": False,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.superadmins.insert_one(super_admin)
    
    return {
        "status": "success",
        "message": "Database seeded successfully",
        "superadmin_email": data.superadmin_email,
        "superadmin_id": super_admin["id"]
    }

class ResetPasswordRequest(BaseModel):
    email: EmailStr
    new_password: str

@api_router.post("/seed/reset-password")
async def reset_superadmin_password(data: ResetPasswordRequest):
    """Reset superadmin password. Use from VPS terminal only."""
    admin = await db.superadmins.find_one({"email": data.email}, {"_id": 0})
    if not admin:
        return {"status": "error", "message": f"Superadmin with email {data.email} not found"}
    
    hashed = hash_password(data.new_password)
    await db.superadmins.update_one(
        {"email": data.email},
        {"$set": {"password": hashed, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"status": "success", "message": f"Password for {data.email} has been reset"}

@api_router.get("/seed/status")
async def seed_status():
    """Check database status - useful for debugging deployment"""
    superadmin_count = await db.superadmins.count_documents({})
    company_count = await db.companies.count_documents({})
    company_admin_count = await db.company_admins.count_documents({})
    employee_count = await db.employees.count_documents({})
    user_count = await db.users.count_documents({})
    
    return {
        "database_name": os.environ.get('DB_NAME', 'unknown'),
        "collections": {
            "superadmins": superadmin_count,
            "companies": company_count,
            "company_admins": company_admin_count,
            "employees": employee_count,
            "users": user_count
        },
        "is_empty": superadmin_count == 0 and company_count == 0,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
